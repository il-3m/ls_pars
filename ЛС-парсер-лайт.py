import sys
import time
import logging
import pandas as pd
import re
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget,
    QLabel, QLineEdit, QPushButton, QCheckBox,
    QDateEdit, QMessageBox, QProgressBar,
    QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QTextEdit,
    QFileDialog, QDialog, QGridLayout, QSpacerItem,
    QSizePolicy, QComboBox, QCompleter, QFormLayout, QStatusBar
)
from PyQt5.QtCore import QDate, Qt, QThread, pyqtSignal, QStringListModel
from PyQt5.QtGui import QColor
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import requests

# Настройка логирования
logging.basicConfig(
    filename='parser_log.txt',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)

MEDICAL_FORMS = [
    "РАСТВОР ДЛЯ ИНФУЗИЙ", "РАСТВОР ДЛЯ ВНУТРИВЕННОГО ВВЕДЕНИЯ",
    "ТАБЛЕТКИ", "ТАБЛЕТКИ, ПОКРЫТЫЕ ОБОЛОЧКОЙ", "КАПСУЛЫ", "СИРОП",
    "СУСПЕНЗИЯ ДЛЯ ПРИЕМА ВНУТРЬ", "МАЗЬ ДЛЯ НАРУЖНОГО ПРИМЕНЕНИЯ",
    "КРЕМ ДЛЯ НАРУЖНОГО ПРИМЕНЕНИЯ", "РАСТВОР ДЛЯ НАРУЖНОГО ПРИМЕНЕНИЯ",
    "КАПЛИ ГЛАЗНЫЕ", "КАПЛИ НАЗАЛЬНЫЕ", "СПРЕЙ НАЗАЛЬНЫЙ",
    "ЛИОФИЛИЗАТ ДЛЯ ПРИГОТОВЛЕНИЯ РАСТВОРА ДЛЯ ИНФУЗИЙ",
    "ПОРОШОК ДЛЯ ПРИГОТОВЛЕНИЯ РАСТВОРА ДЛЯ ИНФУЗИЙ",
    "РАСТВОР ДЛЯ ИНЪЕКЦИЙ", "ГЕЛЬ ДЛЯ НАРУЖНОГО ПРИМЕНЕНИЯ",
    "СУППОЗИТОРИИ РЕКТАЛЬНЫЕ", "КАПЛИ УШНЫЕ", "ТАБЛЕТКИ РАСТВОРИМЫЕ",
    "ТАБЛЕТКИ ДЛЯ РАССАСЫВАНИЯ", "ПОРОШОК ДЛЯ ПРИГОТОВЛЕНИЯ РАСТВОРА ДЛЯ ПРИЕМА ВНУТРЬ"
]


class WorkerThread(QThread):
    update_progress = pyqtSignal(int)
    update_output = pyqtSignal(str)
    finished = pyqtSignal(list)

    def __init__(self, app, search_text, date_from, date_to, moscow_only, rosunimed_only, max_contracts):
        super().__init__()
        self.app = app
        self.search_text = search_text
        self.date_from = date_from
        self.date_to = date_to
        self.moscow_only = moscow_only
        self.rosunimed_only = rosunimed_only
        self.max_contracts = max_contracts
        self.driver = None

    def run(self):
        try:
            self.parse_data()
        except Exception as e:
            self.update_output.emit(f"Ошибка: {str(e)}")
            logging.error(f"Ошибка в потоке: {str(e)}", exc_info=True)
        finally:
            if self.driver:
                self.driver.quit()
            self.finished.emit(self.app.all_results)

    def parse_data(self):
        base_url = "https://zakupki.gov.ru/epz/contract/search/results.html"
        params = {
            "searchString": self.search_text,
            "morphology": "on",
            "search-filter": "Дате+размещения",
            "fz44": "on",
            "contractStageList_1": "on",
            "contractStageData": "1",
            "budgetLevelsIdNameHidden": "{}",
            "contractDateFrom": self.date_from,
            "contractDateTo": self.date_to,
            "sortBy": "UPDATE_DATE",
            "pageNumber": "1",
            "sortDirection": "false",
            "recordsPerPage": "_10",
            "showLotsInfoHidden": "false",
            "strictEqual": "true"
        }

        if self.rosunimed_only:
            params["customerIdOrg"] = '14269:ФЕДЕРАЛЬНОЕ ГОСУДАРСТВЕННОЕ БЮДЖЕТНОЕ ОБРАЗОВАТЕЛЬНОЕ УЧРЕЖДЕНИЕ ВЫСШЕГО ОБРАЗОВАНИЯ "РОССИЙСКИЙ УНИВЕРСИТЕТ МЕДИЦИНЫ" МИНИСТЕРСТВА ЗДРАВООХРАНЕНИЯ РОССИЙСКОЙ ФЕДЕРАЦИИzZ03731000459zZ666998zZ63203zZ7707082145zZ'
        elif self.moscow_only:
            params["customerPlace"] = "77000000000,50000000000"
            params["customerPlaceCodes"] = "77000000000,50000000000"

        url = base_url + "?" + "&".join([f"{k}={v}" for k, v in params.items()])
        self.update_output.emit(f"Запрос: {url}")
        self.update_progress.emit(10)

        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--disable-gcm")
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
        self.driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=chrome_options
        )

        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                self.driver.get(url)
                self.update_output.emit("Ожидание загрузки страницы...")
                self.update_progress.emit(20)
                WebDriverWait(self.driver, 15).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a[href]"))
                )
                break
            except Exception as e:
                self.update_output.emit(f"Попытка {attempt + 1}/{max_attempts} не удалась: {str(e)}")
                if attempt == max_attempts - 1:
                    raise Exception("Не удалось загрузить страницу")
                time.sleep(3)

        try:
            pagination = self.driver.find_elements(By.CSS_SELECTOR, ".paginator a")
            page_numbers = []
            for a in pagination:
                try:
                    page_numbers.append(int(a.text))
                except ValueError:
                    continue
            total_pages = max(page_numbers) if page_numbers else 1
        except Exception:
            total_pages = 1

        all_results = []
        contracts_count = 0

        for page in range(1, total_pages + 1):
            if contracts_count >= self.max_contracts:
                break

            params["pageNumber"] = str(page)
            url = base_url + "?" + "&".join([f"{k}={v}" for k, v in params.items()])
            self.driver.get(url)
            self.update_output.emit(f"Страница {page}/{total_pages}")
            self.update_progress.emit(20 + (page * 60 // total_pages))

            links = self.driver.find_elements(By.CSS_SELECTOR, "a[href]")
            original_links = [
                link.get_attribute("href")
                for link in links
                if link.get_attribute("href") and "contract/contractCard/common-info.html" in link.get_attribute("href")
            ]
            unique_links = list(set(original_links))

            for i, original_link in enumerate(unique_links, 1):
                if contracts_count >= self.max_contracts:
                    break

                self.driver.get(original_link)
                
                if "captcha" in self.driver.page_source.lower():
                    reply = QMessageBox.question(
                        self.app, "CAPTCHA", "Обнаружена CAPTCHA. Продолжить?",
                        QMessageBox.Yes | QMessageBox.No
                    )
                    if reply == QMessageBox.Yes:
                        time.sleep(10)
                    else:
                        continue

                contract_date = self.app.extract_contract_date(self.driver)
                payment_link = original_link.replace("common-info.html", "payment-info-and-target-of-order.html")
                self.driver.get(payment_link)

                if "captcha" in self.driver.page_source.lower():
                    continue

                results = self.app.parse_contract_page(self.driver, self.search_text, contract_date)
                if results:
                    for result in results:
                        if contracts_count >= self.max_contracts:
                            break
                        all_results.append((*result, original_link))
                        contracts_count += 1

        if all_results:
            self.update_output.emit(f"Найдено записей: {len(all_results)}")
            for result in all_results:
                self.app.all_results.append(result)
        else:
            self.update_output.emit("Данные не найдены.")

        self.update_progress.emit(100)


class LogDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Лог парсера")
        self.resize(800, 600)
        self.setModal(False)
        layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)
        self.setLayout(layout)
        self.log_history = []

    def append_log(self, text):
        self.log_history.append(text)
        self.log_text.append(text)


class ZakupkiParserApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.thread = None
        self.all_results = []
        self.filtered_results = []
        self.log_dialog = None
        self.mnn_list = []
        self.mnn_model = QStringListModel()

    def init_ui(self):
        self.setWindowTitle('Парсер лекарств с ЕИС')
        self.setGeometry(100, 100, 1400, 900)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(20, 20, 20, 20)

        self.setStyleSheet("""
            QMainWindow { background-color: #f5f5f5; }
            QLabel { font-size: 14px; font-weight: bold; color: #333; }
            QLineEdit, QDateEdit, QComboBox { 
                font-size: 14px; padding: 8px; 
                border: 1px solid #ccc; border-radius: 5px; 
            }
            QPushButton { 
                font-size: 14px; padding: 10px; 
                background-color: #0078d4; color: white; 
                border: none; border-radius: 5px; 
            }
            QPushButton:hover { background-color: #005ea2; }
            QPushButton:disabled { background-color: #cccccc; }
            QCheckBox { font-size: 14px; color: #333; }
            QProgressBar { 
                border: 1px solid #ccc; border-radius: 5px; 
                text-align: center; font-size: 12px; 
            }
            QTableWidget { 
                gridline-color: #ccc; font-size: 12px;
                selection-background-color: #0078d4;
                selection-color: white;
            }
            QTableWidget::item { padding: 4px; border: 1px solid #eee; }
            QHeaderView::section { 
                font-weight: bold; font-size: 13px; 
                padding: 6px; background-color: #e0e0e0;
            }
        """)

        # Параметры поиска
        params_layout = QFormLayout()
        
        self.search_input = QComboBox()
        self.search_input.setEditable(True)
        self.search_input.setPlaceholderText('Например: АЗИТРОМИЦИН')
        self.search_input.completer().setCompletionMode(QCompleter.PopupCompletion)
        self.search_input.setInsertPolicy(QComboBox.NoInsert)

        self.date_from = QDateEdit()
        self.date_from.setDate(QDate.currentDate().addMonths(-3))
        self.date_from.setCalendarPopup(True)
        
        self.date_to = QDateEdit()
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setCalendarPopup(True)

        self.max_contracts_input = QLineEdit()
        self.max_contracts_input.setText("20")
        self.max_contracts_input.setFixedWidth(60)

        params_layout.addRow(QLabel('Поисковый запрос (МНН):'), self.search_input)
        params_layout.addRow(QLabel('Дата с:'), self.date_from)
        params_layout.addRow(QLabel('Дата по:'), self.date_to)
        params_layout.addRow(QLabel('Макс. контрактов:'), self.max_contracts_input)

        # Фильтры
        filter_layout = QHBoxLayout()
        self.region_checkbox = QCheckBox('Искать только в Москве и МО')
        self.rosunimed_checkbox = QCheckBox('Искать только в Росунимеде')
        self.region_checkbox.toggled.connect(self.on_region_checkbox_toggled)
        self.rosunimed_checkbox.toggled.connect(self.on_rosunimed_checkbox_toggled)
        filter_layout.addWidget(self.region_checkbox)
        filter_layout.addWidget(self.rosunimed_checkbox)
        filter_layout.addStretch()

        # Кнопки
        button_layout = QHBoxLayout()
        self.load_db_button = QPushButton('Загрузить базу МНН')
        self.load_db_button.clicked.connect(self.load_database)
        self.search_button = QPushButton('Найти закупки')
        self.search_button.clicked.connect(self.start_parsing)
        self.export_button = QPushButton('Экспорт в Excel')
        self.export_button.setEnabled(False)
        self.export_button.clicked.connect(self.export_to_excel)
        
        button_layout.addWidget(self.load_db_button)
        button_layout.addWidget(self.search_button)
        button_layout.addWidget(self.export_button)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)

        # Фильтр результатов по МНН (ГРЛС)
        filter_mnn_layout = QHBoxLayout()
        self.filter_mnn_label = QLabel('Фильтр по МНН:')
        self.filter_input = QLineEdit()
        self.filter_input.textChanged.connect(self.filter_results)
        self.filter_input.setPlaceholderText('Введите наименование МНН для фильтрации...')
        filter_mnn_layout.addWidget(self.filter_mnn_label)
        filter_mnn_layout.addWidget(self.filter_input, 1)

        # Таблица результатов
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(17)
        self.results_table.setHorizontalHeaderLabels([
            "Наименование", "Страна", "КТРУ/ОКПД2", "Тип", "Количество", 
            "Цена", "Сумма", "Заказчик", "№ контракта", "№ реестра", 
            "Лек. форма", "Дозировка", "Дата контракта", "Торговое наименование", 
            "Номер РУ", "МНН (ГРЛС)", "Ссылка"
        ])
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.results_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.results_table.itemDoubleClicked.connect(self.show_detail_dialog)

        widths = [250, 100, 120, 80, 100, 100, 120, 200, 120, 120, 120, 100, 100, 150, 120, 150, 100]
        for i, w in enumerate(widths):
            self.results_table.setColumnWidth(i, w)

        # Статистика
        self.stats_layout = QHBoxLayout()
        self.total_purchases_label = QLabel("Всего: 0")
        self.max_price_label = QLabel("Макс: 0")
        self.min_price_label = QLabel("Мин: 0")
        self.avg_price_label = QLabel("Средняя: 0")
        self.stats_layout.addWidget(self.total_purchases_label)
        self.stats_layout.addWidget(self.max_price_label)
        self.stats_layout.addWidget(self.min_price_label)
        self.stats_layout.addWidget(self.avg_price_label)
        self.stats_layout.addStretch()

        # Основной макет
        main_layout.addLayout(params_layout)
        main_layout.addLayout(filter_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.progress_bar)
        main_layout.addLayout(filter_mnn_layout)
        main_layout.addWidget(self.results_table)
        main_layout.addLayout(self.stats_layout)

        central_widget.setLayout(main_layout)

        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.status_label = QLabel("Готов")
        self.statusBar.addPermanentWidget(self.status_label)

    def on_region_checkbox_toggled(self, checked):
        if checked and self.rosunimed_checkbox.isChecked():
            self.rosunimed_checkbox.setChecked(False)

    def on_rosunimed_checkbox_toggled(self, checked):
        if checked and self.region_checkbox.isChecked():
            self.region_checkbox.setChecked(False)

    def check_internet(self):
        try:
            requests.get("https://www.google.com", timeout=5)
            return True
        except:
            return False

    def load_database(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл базы данных ЕСКЛП", "", "Excel Files (*.xlsx *.xls)"
        )
        if not file_path:
            return
        try:
            xl_file = pd.ExcelFile(file_path)
            smnn_sheet_name = None
            for sheet_name in xl_file.sheet_names:
                if sheet_name.startswith("esklp_smnn"):
                    smnn_sheet_name = sheet_name
                    break
            if not smnn_sheet_name:
                QMessageBox.warning(self, 'Ошибка', 'Лист "esklp_smnn" не найден.')
                return
            df = pd.read_excel(file_path, sheet_name=smnn_sheet_name, header=0)
            mnn_series = df.iloc[:, 0].dropna().unique()
            self.mnn_list = [str(mnn).strip() for mnn in mnn_series if str(mnn).strip()]
            self.mnn_model.setStringList(self.mnn_list)
            self.search_input.setModel(self.mnn_model)
            self.search_input.completer().setModel(self.mnn_model)
            QMessageBox.information(self, 'Успех', f'Загружено {len(self.mnn_list)} МНН.')
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Не удалось загрузить: {str(e)}')

    def extract_contract_date(self, driver):
        try:
            title_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "//span[@class='section__title' and contains(text(), 'Дата заключения контракта')]")
                )
            )
            info_element = title_element.find_element(By.XPATH, "./following-sibling::span[@class='section__info']")
            if info_element:
                text = info_element.text.strip()
                date_match = re.search(r'\b\d{2}\.\d{2}\.\d{4}\b', text)
                if date_match:
                    return date_match.group(0)
        except Exception as e:
            logging.error(f"Ошибка при извлечении даты: {e}")
        return "Не указано"

    def expand_medical_details(self, driver, search_text):
        """Принудительно раскрывает все блоки с помощью JavaScript"""
        try:
            logging.info(f"  === Раскрытие блоков для '{search_text}' ===")
            
            expanded_count = 0
            
            # Способ 1: Кликаем по всем toggle-элементам
            toggle_selectors = [
                "button[class*='toggle']",
                "[class*='expand']",
                "[class*='collapse']",
                ".purchase-object__header",
                ".lot-info__toggle",
                "[aria-expanded='false']",
            ]
            
            for selector in toggle_selectors:
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    for el in elements:
                        try:
                            aria_expanded = el.get_attribute("aria-expanded")
                            if aria_expanded == "true":
                                continue
                            
                            driver.execute_script("""
                                arguments[0].scrollIntoView({behavior: 'auto', block: 'center'});
                                arguments[0].click();
                            """, el)
                            
                            expanded_count += 1
                        except:
                            continue
                except:
                    continue
            
            # Способ 2: Принудительно показываем скрытые элементы
            try:
                driver.execute_script("""
                    var hiddenElements = document.querySelectorAll('[style*="display: none"], [style*="visibility: hidden"]');
                    hiddenElements.forEach(function(el) {
                        el.style.display = 'block';
                        el.style.visibility = 'visible';
                        el.style.opacity = '1';
                    });
                    
                    var collapsedSections = document.querySelectorAll('.collapse, [class*="collapse"]');
                    collapsedSections.forEach(function(el) {
                        el.classList.remove('collapse');
                        el.classList.add('show');
                        el.style.display = 'block';
                    });
                    
                    var expandableElements = document.querySelectorAll('[aria-expanded]');
                    expandableElements.forEach(function(el) {
                        el.setAttribute('aria-expanded', 'true');
                    });
                """)
                logging.info("    Принудительно показаны скрытые элементы")
            except Exception as e:
                logging.debug(f"    Ошибка при показе: {e}")
            
            # Ждем появления таблиц
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
                )
            except:
                pass
            
            time.sleep(2)
            logging.info(f"  === Обработано блоков: {expanded_count} ===")
            
        except Exception as e:
            logging.error(f"  ⚠ Ошибка раскрытия: {e}")

    def parse_contract_page(self, driver, search_text, contract_date):
        """Парсинг страницы контракта с раскрытием всех блоков"""
        try:
            # Раскрываем ВСЕ блоки
            self.expand_medical_details(driver, search_text)
            
            WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "table"))
            )
            
            all_tables = driver.find_elements(By.CSS_SELECTOR, "table")
            logging.info(f"  Найдено таблиц: {len(all_tables)}")
            
            results = []
            search_text_lower = search_text.lower()
            
            for table_num, table in enumerate(all_tables, 1):
                try:
                    rows = table.find_elements(By.CSS_SELECTOR, "tr")
                    if len(rows) < 2:
                        continue
                    
                    table_text = table.text.lower()
                    
                    if search_text_lower not in table_text:
                        continue
                    
                    logging.info(f"  Таблица {table_num} содержит '{search_text}'")
                    
                    header_indices = {
                        "наименование": None, "ктру": None, "тип": None,
                        "количество": None, "цена": None, "сумма": None
                    }
                    
                    header_row = rows[0]
                    header_cells = header_row.find_elements(By.CSS_SELECTOR, "th, td")
                    header_texts = [cell.text.strip().lower() for cell in header_cells]
                    
                    for i, text in enumerate(header_texts):
                        if "наименование" in text or "объект закупки" in text:
                            header_indices["наименование"] = i
                        elif "ктру" in text or "окпд" in text:
                            header_indices["ктру"] = i
                        elif "тип" in text:
                            header_indices["тип"] = i
                        elif "количество" in text or "объем" in text:
                            header_indices["количество"] = i
                        elif "цена" in text:
                            header_indices["цена"] = i
                        elif "сумма" in text:
                            header_indices["сумма"] = i
                    
                    # Таблица с деталями лекарств
                    if any(h in table_text for h in ["торговое наименование", "номер ру", "лекарственная форма", "дозировка"]):
                        logging.info(f"    Таблица с деталями лекарств!")
                        
                        tn_idx = ru_idx = form_idx = dose_idx = None
                        
                        for idx, h in enumerate(header_texts):
                            if "торговое наименование" in h or "тн" in h:
                                tn_idx = idx
                            elif "номер ру" in h or "ру" in h:
                                ru_idx = idx
                            elif "лекарственная форма" in h or "форма" in h:
                                form_idx = idx
                            elif "дозировка" in h or "доза" in h:
                                dose_idx = idx
                        
                        for row in rows[1:]:
                            cells = row.find_elements(By.CSS_SELECTOR, "td")
                            if not cells:
                                continue
                            
                            # Используем textContent вместо .text, чтобы захватить весь текст из всех span-элементов
                            cell_texts = []
                            for c in cells:
                                full_text = c.get_attribute("textContent")
                                if full_text:
                                    # Удаляем лишние пробелы и переносы строк, но сохраняем содержимое всех span
                                    full_text = ' '.join(full_text.split())
                                cell_texts.append(full_text.strip() if full_text else "")
                            
                            if any(search_text_lower in c.lower() for c in cell_texts):
                                trade_name = reg_number = medical_form = dosage = "Не указано"
                                
                                if tn_idx is not None and tn_idx < len(cell_texts) and cell_texts[tn_idx]:
                                    trade_name = cell_texts[tn_idx]
                                if ru_idx is not None and ru_idx < len(cell_texts) and cell_texts[ru_idx]:
                                    reg_number = cell_texts[ru_idx]
                                if form_idx is not None and form_idx < len(cell_texts) and cell_texts[form_idx]:
                                    form_text = cell_texts[form_idx].upper()
                                    for form in sorted(MEDICAL_FORMS, key=len, reverse=True):
                                        if form in form_text:
                                            medical_form = form
                                            break
                                if dose_idx is not None and dose_idx < len(cell_texts) and cell_texts[dose_idx]:
                                    dosage = cell_texts[dose_idx]
                                
                                results.append((
                                    search_text, "Не указано", "Не указано", "Товар",
                                    "Не указано", "Не указано", "Не указано",
                                    "Не указано", "Не указано", "Не указано",
                                    medical_form, dosage, contract_date,
                                    trade_name, reg_number, search_text
                                ))
                                
                                logging.info(f"    ТН={trade_name}, РУ={reg_number}, Форма={medical_form}, Доза={dosage}")
                        continue
                    
                    if header_indices["наименование"] is None:
                        continue
                    
                    for row in rows[1:]:
                        cells = row.find_elements(By.CSS_SELECTOR, "td")
                        if not cells:
                            continue
                        
                        cell_texts = [cell.text.strip() for cell in cells]
                        
                        found = any(search_text_lower in c.lower() for c in cell_texts)
                        if not found:
                            continue
                        
                        def get_cell_text(index):
                            return cell_texts[index] if index is not None and index < len(cell_texts) else "Не указано"
                        
                        product_name = get_cell_text(header_indices["наименование"])
                        ktu_okpd = get_cell_text(header_indices["ктру"])
                        type_object = get_cell_text(header_indices["тип"]) or "Товар"
                        quantity_unit = get_cell_text(header_indices["количество"])
                        unit_price = get_cell_text(header_indices["цена"])
                        total_price_vat = get_cell_text(header_indices["сумма"])
                        
                        if product_name.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.", "9.")):
                            product_name = product_name[product_name.find(".") + 1:].strip()
                        
                        country_text = "Не указано"
                        country_match = re.search(r'Страна происхождения:\s*([^,\n]+)', product_name)
                        if country_match:
                            country_text = country_match.group(1).strip()
                            product_name = re.sub(r'Страна происхождения:\s*[^,\n]+', '', product_name).strip()
                        
                        product_name = re.sub(r'единица измерения товара: Штука \(шт\)', '', product_name).strip()
                        
                        if unit_price and unit_price != "Не указано":
                            price_match = re.search(r'([\d\s.,]+)', unit_price.replace(' ', ''))
                            unit_price = price_match.group(1).replace(' ', '') if price_match else "Не указано"
                        
                        if total_price_vat and total_price_vat != "Не указано":
                            sum_match = re.search(r'([\d\s.,]+)', total_price_vat.replace(' ', ''))
                            sum_value = sum_match.group(1).replace(' ', '') if sum_match else ""
                            nds_match = re.search(r'(НДС.*?\d+%)', total_price_vat, re.IGNORECASE)
                            nds_value = nds_match.group(1) if nds_match else ""
                            if sum_value and nds_value:
                                total_price_vat = f"{sum_value} ({nds_value})"
                            elif sum_value:
                                total_price_vat = sum_value
                        
                        customer = contract_number = reestr_number = "Не указано"
                        
                        current_url = driver.current_url
                        reestr_match = re.search(r'reestrNumber=([0-9]+)', current_url)
                        if reestr_match:
                            reestr_number = reestr_match.group(1)
                        
                        try:
                            sections = driver.find_elements(By.CLASS_NAME, "cardMainInfo__section")
                            for section in sections:
                                section_text = section.text.strip()
                                if "Заказчик" in section_text and customer == "Не указано":
                                    lines = section_text.split('\n')
                                    for i, line in enumerate(lines):
                                        if "Заказчик" in line and i + 1 < len(lines):
                                            potential_customer = lines[i + 1].strip()
                                            if not potential_customer.startswith(("Заказчик:", "Контракт:")):
                                                customer = potential_customer
                                                break
                                if "Контракт" in section_text and contract_number == "Не указано":
                                    contract_match = re.search(r'Контракт[^\d№]*[№\s]*([^,\n]+)', section_text)
                                    if contract_match:
                                        contract_number = contract_match.group(1).strip()
                                        if contract_number.startswith("№"):
                                            contract_number = contract_number[1:].strip()
                                        break
                        except:
                            pass
                        
                        medical_form = dosage = "Не указано"
                        sorted_forms = sorted(MEDICAL_FORMS, key=len, reverse=True)
                        for form in sorted_forms:
                            if form in product_name.upper():
                                medical_form = form
                                break
                        
                        dosage_pattern = r'(\d+(?:[.,]\d+)?\s*(?:мг|мл|ед|мкг|г))'
                        dosage_matches = re.findall(dosage_pattern, product_name, re.IGNORECASE)
                        if dosage_matches:
                            dosage = ', '.join(dosage_matches)
                        
                        results.append((
                            product_name, country_text, ktu_okpd, type_object,
                            quantity_unit, unit_price, total_price_vat,
                            customer, contract_number, reestr_number,
                            medical_form, dosage, contract_date,
                            "Не указано", "Не указано", search_text
                        ))
                        
                except Exception as e:
                    logging.debug(f"  Ошибка таблицы {table_num}: {e}")
                    continue
            
            logging.info(f"  === Извлечено записей: {len(results)} ===")
            return results
            
        except Exception as e:
            logging.error(f"Ошибка парсинга: {str(e)}", exc_info=True)
            return []

    def start_parsing(self):
        if not self.check_internet():
            QMessageBox.warning(self, 'Ошибка', 'Нет интернет-соединения')
            return

        search_text = self.search_input.currentText().strip()
        if not search_text:
            QMessageBox.warning(self, 'Ошибка', 'Введите поисковый запрос')
            return

        self.search_button.setEnabled(False)
        self.export_button.setEnabled(False)
        self.progress_bar.setValue(0)
        self.results_table.setRowCount(0)
        self.all_results = []
        self.filtered_results = []

        date_from = self.date_from.date().toString('dd.MM.yyyy')
        date_to = self.date_to.date().toString('dd.MM.yyyy')
        moscow_only = self.region_checkbox.isChecked()
        rosunimed_only = self.rosunimed_checkbox.isChecked()

        try:
            max_contracts = int(self.max_contracts_input.text())
            if max_contracts <= 0:
                raise ValueError()
        except:
            QMessageBox.warning(self, 'Ошибка', 'Введите корректное число')
            self.search_button.setEnabled(True)
            return

        self.thread = WorkerThread(self, search_text, date_from, date_to, moscow_only, rosunimed_only, max_contracts)
        self.thread.update_progress.connect(self.progress_bar.setValue)
        self.thread.update_output.connect(self.on_log_update)
        self.thread.finished.connect(self.on_parsing_finished)
        self.thread.start()

        if self.log_dialog is None:
            self.log_dialog = LogDialog(self)
            self.log_dialog.show()
        self.log_dialog.append_log(f"Начат поиск: {search_text}")

    def on_log_update(self, text):
        if self.log_dialog:
            self.log_dialog.append_log(text)

    def on_parsing_finished(self, results):
        self.search_button.setEnabled(True)
        self.export_button.setEnabled(bool(results))
        self.progress_bar.setValue(100)
        self.filtered_results = self.all_results.copy()
        self.display_results()
        self.update_stats()
        self.status_label.setText(f"Готов (найдено: {len(results)})")

    def display_results(self):
        self.results_table.setRowCount(0)
        for row_data in self.filtered_results:
            row_position = self.results_table.rowCount()
            self.results_table.insertRow(row_position)
            for i in range(min(16, len(row_data))):
                item = QTableWidgetItem(str(row_data[i]) if i < len(row_data) else "")
                item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                self.results_table.setItem(row_position, i, item)

            link = row_data[-1] if len(row_data) > 16 else ""
            link_item = QTableWidgetItem("Открыть")
            link_item.setData(Qt.UserRole, link)
            link_item.setFlags(link_item.flags() ^ Qt.ItemIsEditable)
            self.results_table.setItem(row_position, 16, link_item)
            self.results_table.setRowHeight(row_position, 60)

    def filter_results(self):
        filter_text = self.filter_input.text().lower()
        # Если фильтр содержит текст, фильтруем по колонке "МНН (ГРЛС)" (индекс 15)
        if filter_text.strip():
            self.filtered_results = [
                result for result in self.all_results
                if filter_text.lower() in str(result[15]).lower()  # Индекс 15 - это "МНН (ГРЛС)"
            ]
        else:
            self.filtered_results = self.all_results.copy()
        self.display_results()
        self.update_stats()

    def update_stats(self):
        if not self.filtered_results:
            self.total_purchases_label.setText("Всего: 0")
            self.max_price_label.setText("Макс: 0")
            self.min_price_label.setText("Мин: 0")
            self.avg_price_label.setText("Средняя: 0")
            return

        prices = []
        for row_data in self.filtered_results:
            price_str = row_data[5]
            if price_str and price_str != "Не указано":
                try:
                    price = float(str(price_str).replace(' ', '').replace(',', '.'))
                    prices.append(price)
                except:
                    pass

        if not prices:
            self.total_purchases_label.setText(f"Всего: {len(self.filtered_results)}")
            self.max_price_label.setText("Макс: 0")
            self.min_price_label.setText("Мин: 0")
            self.avg_price_label.setText("Средняя: 0")
            return

        self.total_purchases_label.setText(f"Всего: {len(self.filtered_results)}")
        self.max_price_label.setText(f"Макс: {max(prices):.2f}")
        self.min_price_label.setText(f"Мин: {min(prices):.2f}")
        self.avg_price_label.setText(f"Средняя: {sum(prices)/len(prices):.2f}")

    def export_to_excel(self):
        if not self.all_results:
            QMessageBox.warning(self, 'Ошибка', 'Нет данных')
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить", "results.xlsx", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        df = pd.DataFrame(self.all_results, columns=[
            "Наименование", "Страна", "КТРУ/ОКПД2", "Тип", "Количество",
            "Цена", "Сумма", "Заказчик", "№ контракта", "№ реестра",
            "Лек. форма", "Дозировка", "Дата контракта", "Торговое наименование",
            "Номер РУ", "МНН (ГРЛС)", "Ссылка"
        ])

        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Результаты')
            workbook = writer.book
            worksheet = writer.sheets['Результаты']

            for col_idx, column in enumerate(df.columns, 1):
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 20
                for cell in worksheet[f'{get_column_letter(col_idx)}']:
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    cell.font = Font(size=10)

            for row_idx, row in df.iterrows():
                link = row['Ссылка']
                if pd.notna(link):
                    cell = worksheet.cell(row=row_idx + 2, column=16)
                    cell.hyperlink = link
                    cell.style = 'Hyperlink'

        QMessageBox.information(self, 'Экспорт', f'Файл сохранен: {file_path}')

    def show_detail_dialog(self, item):
        row = item.row()
        if row < 0 or row >= len(self.filtered_results):
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Детали закупки")
        dialog.resize(700, 600)
        layout = QGridLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)

        row_data = self.filtered_results[row]

        fields = [
            ("Наименование", row_data[0]), ("Страна", row_data[1]),
            ("КТРУ/ОКПД2", row_data[2]), ("Тип", row_data[3]),
            ("Количество", row_data[4]), ("Цена", row_data[5]),
            ("Сумма", row_data[6]), ("Заказчик", row_data[7]),
            ("№ контракта", row_data[8]), ("№ реестра", row_data[9]),
            ("Лек. форма", row_data[10]), ("Дозировка", row_data[11]),
            ("Дата контракта", row_data[12]), ("Торговое наименование", row_data[13]),
            ("Номер РУ", row_data[14]), ("МНН (ГРЛС)", row_data[15]),
            ("Ссылка", row_data[16] if len(row_data) > 16 else "")
        ]

        for i, (label, value) in enumerate(fields):
            label_widget = QLabel(label)
            label_widget.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            label_widget.setMinimumWidth(200)
            label_widget.setStyleSheet("font-weight: bold;")

            value_widget = QLabel(str(value))
            value_widget.setWordWrap(True)
            value_widget.setMinimumWidth(350)
            if str(value).startswith("http"):
                value_widget.setText(f'<a href="{value}">{value}</a>')
                value_widget.setOpenExternalLinks(True)
                value_widget.setStyleSheet("color: blue; text-decoration: underline;")

            layout.addWidget(label_widget, i, 0)
            layout.addWidget(value_widget, i, 1)

        dialog.setLayout(layout)
        dialog.exec_()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ZakupkiParserApp()
    window.show()
    sys.exit(app.exec_())