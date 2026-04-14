#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Production parser for EIS contract pages ("Объекты закупки").

Key properties:
- Uses Playwright for stable DOM interaction
- Parses by table headers instead of brittle fixed indices
- Saves debug archive (raw HTML, parse JSON, screenshots)
- Supports single URL and batch mode
- Exports CSV always, XLSX optionally (if pandas/openpyxl installed)
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import json
import re
import sys
import threading
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

import tkinter as tk
from tkinter import ttk, messagebox

from playwright.async_api import Browser, BrowserContext, Error, Frame, Page, async_playwright


FIELD_ORDER = [
    "name",
    "category_ls",
    "okpd2",
    "country",
    "mnn",
    "trade_name",
    "ru",
    "release_form",
    "dose",
    "qty_consumption_unit",
    "price_per_unit",
    "sum_rub",
    "nds",
    "holder_name",
    "manufacturer_name",
    "manufacturer_country",
    "primary_package_type",
    "qty_forms_primary",
    "qty_primary_packages",
    "qty_consumer_units",
    "consumer_package_completeness",
    "contract_date",
    "contract_number",
    "reestr_number",
    "customer_name",
    "contract_link",
]

EXPORT_HEADERS_RU = {
    "name": "Наименование",
    "category_ls": "Категории ЛС",
    "okpd2": "ОКПД2",
    "country": "Страна происхождения",
    "mnn": "МНН (ГРЛС)",
    "trade_name": "Торговое наименование",
    "ru": "РУ",
    "release_form": "Форма выпуска",
    "dose": "Дозировка",
    "qty_consumption_unit": "Количество в потреб. единице измерения",
    "price_per_unit": "Цена за единицу измерения, ₽",
    "sum_rub": "Сумма, ₽",
    "nds": "НДС",
    "holder_name": "Наименование держателя или владельца РУ",
    "manufacturer_name": "Наименование производителя",
    "manufacturer_country": "Страна производителя",
    "primary_package_type": "Вид первичной упаковки",
    "qty_forms_primary": "Количество лекарственных форм в первичной упаковке",
    "qty_primary_packages": "Количество первичных упаковок в потребительской упаковке",
    "qty_consumer_units": "Количество потребительских единиц в потребительской упаковке",
    "consumer_package_completeness": "Комплектность потребительской упаковки",
    "contract_date": "Дата контракта",
    "contract_number": "Номер контракта",
    "reestr_number": "Номер реестра",
    "customer_name": "Наименование заказчика",
    "contract_link": "Ссылка на контракт в ЕИС",
}


@dataclass
class ParseRecord:
    name: str = ""
    category_ls: str = ""
    okpd2: str = ""
    country: str = ""
    mnn: str = ""
    trade_name: str = ""
    ru: str = ""
    release_form: str = ""
    dose: str = ""
    qty_consumption_unit: str = ""
    price_per_unit: str = ""
    sum_rub: str = ""
    nds: str = ""
    holder_name: str = ""
    manufacturer_name: str = ""
    manufacturer_country: str = ""
    primary_package_type: str = ""
    qty_forms_primary: str = ""
    qty_primary_packages: str = ""
    qty_consumer_units: str = ""
    consumer_package_completeness: str = ""
    contract_date: str = ""
    contract_number: str = ""
    reestr_number: str = ""
    customer_name: str = ""
    contract_link: str = ""

    def as_row(self) -> dict[str, str]:
        return {k: getattr(self, k, "") for k in FIELD_ORDER}


class EISParser:
    def __init__(self, timeout_ms: int = 60000, expand_rounds: int = 4, page_load_delay: int = 1200, expand_delay: int = 800) -> None:
        self.timeout_ms = timeout_ms
        self.expand_rounds = expand_rounds
        self.page_load_delay = page_load_delay
        self.expand_delay = expand_delay

    async def parse_url(
        self,
        page: Page,
        url: str,
        archive_dir: Path,
        save_trace: bool = False,
        common_info_url: Optional[str] = None,
    ) -> list[dict[str, str]]:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        job_dir = archive_dir / f"job_{ts}"
        raw_dir = job_dir / "raw"
        parsed_dir = job_dir / "parsed"
        raw_dir.mkdir(parents=True, exist_ok=True)
        parsed_dir.mkdir(parents=True, exist_ok=True)

        # Извлечение метаданных контракта
        contract_date = ""
        contract_number = ""
        reestr_number = ""
        customer_name = ""

        # 1. Извлекаем номер реестра из URL
        reestr_match = re.search(r'reestrNumber=([0-9]+)', url)
        if reestr_match:
            reestr_number = reestr_match.group(1)

        # 2. Если есть common_info_url, парсим дату контракта
        if common_info_url:
            try:
                await page.goto(common_info_url, wait_until="domcontentloaded", timeout=self.timeout_ms)
                await page.wait_for_timeout(self.page_load_delay)
                await self._close_overlays(page)
                
                # XPath для даты заключения контракта
                contract_date = await page.evaluate('''() => {
                    const xpath = "//span[@class='section__title' and contains(text(), 'Дата заключения контракта')]/following-sibling::span[@class='section__info']";
                    const result = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
                    if (result.singleNodeValue) {
                        const text = result.singleNodeValue.textContent || "";
                        const match = text.match(/\\b\\d{2}\\.\\d{2}\\.\\d{4}\\b/);
                        return match ? match[0] : "Не указано";
                    }
                    return "Не указано";
                }''')
            except Exception as e:
                contract_date = "Не указано"
                logging.warning(f"Не удалось извлечь дату контракта из {common_info_url}: {e}")

        # 3. Переходим на основную страницу для парсинга таблицы и метаданных
        await page.goto(url, wait_until="domcontentloaded", timeout=self.timeout_ms)
        await page.wait_for_timeout(self.page_load_delay)

        # 4. Извлекаем номер контракта и заказчика из payment-info-and-target-of-order.html
        try:
            await self._close_overlays(page)
            metadata = await page.evaluate('''() => {
                const blocks = document.querySelectorAll('.cardMainInfo__section');
                let contractNumber = "";
                let customerName = "";
                
                for (const block of blocks) {
                    const text = block.textContent || "";
                    
                    // Поиск номера контракта - используем улучшенный regex
                    if (!contractNumber && text.includes('Контракт')) {
                        const match = text.match(/Контракт\\s*[:\\s]*№?\\s*([A-Za-zА-Яа-я0-9\\-/]+)/);
                        if (match) {
                            contractNumber = match[1].trim();
                        }
                    }
                    
                    // Поиск заказчика - улучшенная логика
                    if (!customerName && text.includes('Заказчик')) {
                        const lines = text.split('\\n').map(l => l.trim()).filter(l => l.length > 0);
                        for (let i = 0; i < lines.length - 1; i++) {
                            const line = lines[i];
                            // Проверяем, что строка содержит "Заказчик" как заголовок
                            if (line.includes('Заказчик') && (line.endsWith(':') || line.match(/^Заказчик\\s*$/))) {
                                const candidate = lines[i + 1];
                                // Отфильтровываем служебные строки
                                if (candidate && 
                                    !candidate.startsWith('Заказчик:') && 
                                    !candidate.startsWith('Контракт:') && 
                                    !candidate.startsWith('Реестровый номер:') &&
                                    candidate.length > 5) {
                                    customerName = candidate;
                                    break;
                                }
                            }
                        }
                    }
                }
                
                return { contractNumber, customerName };
            }''')
            contract_number = metadata.get('contractNumber', '') or contract_number
            customer_name = metadata.get('customerName', '') or customer_name
        except Exception as e:
            logging.warning(f"Не удалось извлечь метаданные контракта: {e}")
        
        # Резервный способ извлечения заказчика через XPath
        if not customer_name:
            try:
                customer_name = await page.evaluate('''() => {
                    const xpath = "//span[contains(text(), 'Заказчик')]/following::span[1]";
                    const result = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null);
                    if (result.singleNodeValue) {
                        const text = result.singleNodeValue.textContent || "";
                        return text.trim();
                    }
                    return "";
                }''')
            except Exception:
                pass
        
        # Второй резервный способ - ищем в любом месте страницы
        if not customer_name:
            try:
                customer_name = await page.evaluate('''() => {
                    const allText = document.body.innerText;
                    const lines = allText.split('\\n').map(l => l.trim()).filter(l => l.length > 0);
                    
                    for (let i = 0; i < lines.length - 1; i++) {
                        const line = lines[i];
                        // Ищем строку с "Заказчик:" в конце
                        if (line.includes('Заказчик:') || line.match(/^Заказчик\\s*$/)) {
                            const candidate = lines[i + 1];
                            if (candidate && 
                                !candidate.startsWith('Заказчик:') && 
                                !candidate.startsWith('Контракт:') && 
                                !candidate.startsWith('Реестровый номер:') &&
                                candidate.length > 10 &&
                                !candidate.includes('ИНН') && 
                                !candidate.includes('КПП')) {
                                return candidate;
                            }
                        }
                    }
                    return "";
                }''')
            except Exception:
                pass

        await self._close_overlays(page)
        frame = await self._find_frame_with_objects(page)
        if frame is None:
            raise RuntimeError("Не найден контекст с блоком 'Объекты закупки'")

        await self._expand_objects(frame)
        await page.wait_for_timeout(self.expand_delay)

        html = await frame.content()
        (raw_dir / "objects_frame.html").write_text(html, encoding="utf-8")

        payload = await frame.evaluate(EXTRACT_SCRIPT)
        records = [self._normalize_record(item) for item in payload or []]
        records = self._merge_split_records(records)
        records = [self._finalize_record(rec) for rec in records]
        
        # Добавляем метаданные контракта в каждую запись
        for rec in records:
            rec.contract_date = contract_date
            rec.contract_number = contract_number
            rec.reestr_number = reestr_number
            rec.customer_name = customer_name
            rec.contract_link = url  # Сохраняем ссылку на страницу контракта
        
        rows = [r.as_row() for r in records]

        (parsed_dir / "rows.json").write_text(
            json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        if save_trace:
            # trace is started in outer layer (context), this only marks run folder
            (job_dir / "trace_requested.txt").write_text("trace enabled", encoding="utf-8")

        return rows

    async def _close_overlays(self, page: Page) -> None:
        candidates = [
            "button:has-text('Принять')",
            "button:has-text('Согласен')",
            "button:has-text('Закрыть')",
            "[aria-label='close']",
            "[aria-label='Close']",
        ]
        for selector in candidates:
            loc = page.locator(selector)
            count = min(await loc.count(), 3)
            for i in range(count):
                try:
                    await loc.nth(i).click(timeout=800)
                except Error:
                    pass

    async def _find_frame_with_objects(self, page: Page) -> Optional[Frame]:
        for frame in page.frames:
            try:
                if await frame.locator("text=Объекты закупки").count() > 0:
                    return frame
            except Error:
                continue
        return None

    async def _expand_objects(self, frame: Frame) -> None:
        try:
            anchor = frame.locator("text=Объекты закупки").first
            await anchor.scroll_into_view_if_needed(timeout=4000)
        except Error:
            pass

        for _ in range(self.expand_rounds):
            toggles = frame.locator(
                "[aria-expanded='false'], button[aria-expanded='false'], [role='button'][aria-expanded='false']"
            )
            cnt = await toggles.count()
            if cnt == 0:
                break

            clicked = 0
            for i in range(min(cnt, 120)):
                item = toggles.nth(i)
                try:
                    await item.scroll_into_view_if_needed(timeout=1200)
                    await item.click(timeout=1200)
                    clicked += 1
                except Error:
                    continue

            if clicked == 0:
                break
            await frame.page.wait_for_timeout(500)

    def _normalize_record(self, raw: dict) -> ParseRecord:
        rec = ParseRecord(**{k: _clean(raw.get(k, "")) for k in FIELD_ORDER})

        # Stage 2 normalization: keep raw extraction, then reshape columns from mixed blocks.
        blob = " | ".join(_clean(raw.get(k, "")) for k in FIELD_ORDER if _clean(raw.get(k, "")))

        # Name must not be equal to OKPD2 code.
        if _looks_like_okpd2_code(rec.name) or (rec.okpd2 and _clean(rec.name) == _clean(rec.okpd2)):
            rec.name = ""

        # Fallback for legacy payloads where one compact block is stored in country.
        if "|" in rec.country and (not rec.category_ls or not rec.okpd2):
            self._split_compact_top_block(rec)

        # Parse compact pipe blocks from any spilled field (country/qty/name blobs).
        self._split_compact_top_block(rec, source_text=blob)

        # Recover the item name when it leaked into other columns.
        if not rec.name or len(rec.name) < 6:
            parsed_name = _extract_name_from_blob(blob)
            if parsed_name:
                rec.name = parsed_name

        # Fix legacy column shift: RU -> trade_name, form -> RU, dose -> form, qty -> dose.
        if rec.ru and not _looks_like_ru_code(rec.ru) and _looks_like_ru_code(rec.release_form):
            if not rec.trade_name:
                rec.trade_name = rec.ru
            rec.ru = rec.release_form
            rec.release_form = rec.dose
            rec.dose = ""

        # Additional shift case: trade_name is empty, RU stores trade text, release_form stores RU.
        if not rec.trade_name and rec.ru and _looks_like_ru_code(rec.release_form):
            rec.trade_name = rec.ru
            rec.ru = rec.release_form
            if rec.dose and not rec.release_form:
                rec.release_form = rec.dose
            rec.dose = _extract_dose(rec.dose) or ""

        # Pull detailed fields from long GRLS block when it was flattened into one cell.
        if _looks_like_meta_blob(blob):
            meta = _extract_meta_fields(blob)
            for field, value in meta.items():
                if getattr(rec, field, "") == "" and value:
                    setattr(rec, field, value)

            # If qty column contains a long block, keep only real qty text.
            if len(rec.qty_consumption_unit) > 80:
                qty_candidate = _extract_qty_from_blob(blob)
                if qty_candidate:
                    rec.qty_consumption_unit = qty_candidate

        # Sometimes completeness contains a tail with item title and compact top-row values.
        if rec.consumer_package_completeness and (
            re.search(r"\d+\.\s+", rec.consumer_package_completeness)
            or "СТРАНА ПРОИСХОЖДЕНИЯ" in rec.consumer_package_completeness.upper()
        ):
            carried_name = _extract_name_from_blob(rec.consumer_package_completeness)
            if carried_name and (not rec.name or _looks_like_okpd2_code(rec.name)):
                rec.name = carried_name

            rec.consumer_package_completeness = "~"

        if not rec.okpd2:
            rec.okpd2 = _extract_okpd2(rec.name) or _extract_okpd2(rec.category_ls)

        dose_candidates = [_extract_dose(rec.dose), _extract_dose(rec.name), _extract_dose(blob)]
        dose_candidates = [d for d in dose_candidates if d]
        dose_candidate = ""
        if dose_candidates:
            # Prefer complex dosage chains (with '+') and then the longest variant.
            complex_candidates = [d for d in dose_candidates if "+" in d]
            dose_candidate = max(complex_candidates or dose_candidates, key=len)
        if dose_candidate:
            rec.dose = dose_candidate

        if not rec.country:
            rec.country = _extract_country(rec.name)

        if not rec.dose:
            rec.dose = _extract_dose(blob)

        # Quantity/price/sum should come from top compact block, not from metadata text.
        if _looks_like_meta_blob(rec.qty_consumption_unit) or "СТРАНА ПРОИСХОЖДЕНИЯ" in rec.qty_consumption_unit.upper():
            rec.qty_consumption_unit = ""

        if rec.price_per_unit and not _looks_like_price(rec.price_per_unit):
            rec.price_per_unit = ""

        if rec.sum_rub and not _looks_like_sum(rec.sum_rub):
            rec.sum_rub = ""

        compact = _extract_compact_values(blob)
        if not rec.qty_consumption_unit and compact.get("qty_consumption_unit"):
            rec.qty_consumption_unit = compact["qty_consumption_unit"]
        if not rec.price_per_unit and compact.get("price_per_unit"):
            rec.price_per_unit = compact["price_per_unit"]
        if not rec.sum_rub and compact.get("sum_rub"):
            rec.sum_rub = compact["sum_rub"]

        if not rec.qty_consumption_unit:
            rec.qty_consumption_unit = _extract_qty_from_blob(blob)

        if not rec.price_per_unit:
            rec.price_per_unit = _extract_price(blob)

        if not rec.sum_rub:
            rec.sum_rub = _extract_sum(blob)

        rec.country = _short_country(rec.country)

        # Keep name complete when source row ended up in another column.
        if (not rec.name or _looks_like_okpd2_code(rec.name)):
            recovered_name = _extract_name_from_blob(blob)
            if recovered_name:
                rec.name = recovered_name

        # Fallback: compose a readable name from stable columns instead of copying category.
        if not rec.name and rec.trade_name:
            chunks = [rec.trade_name, rec.release_form, rec.dose]
            rec.name = _clean(", ".join([c for c in chunks if c]))

        if _looks_like_okpd2_code(rec.name):
            rec.name = ""

        if not rec.name and rec.trade_name:
            chunks = [rec.trade_name, rec.release_form, rec.dose]
            rec.name = _clean(", ".join([c for c in chunks if c]))

        # If category contains an item-like title, move it to name and keep parsed category.
        if rec.category_ls and (re.match(r"^\d+\.\s", rec.category_ls) or ":" in rec.category_ls):
            if not rec.name or len(rec.name) < len(rec.category_ls):
                rec.name = rec.category_ls
            # Category should stay a class, not a full item title.
            if rec.okpd2:
                rec.category_ls = ""

        if rec.category_ls and rec.name and rec.category_ls == rec.name:
            rec.category_ls = ""

        return rec

    def _finalize_record(self, rec: ParseRecord) -> ParseRecord:
        row_blob = " | ".join(_clean(getattr(rec, f, "")) for f in FIELD_ORDER if _clean(getattr(rec, f, "")))

        if not rec.name or _looks_like_okpd2_code(rec.name):
            rec.name = _extract_name_from_blob(row_blob)
        if (not rec.name or _looks_like_okpd2_code(rec.name)) and rec.trade_name:
            rec.name = _clean(", ".join([x for x in [rec.trade_name, rec.release_form, rec.dose] if _clean(x)])
            )

        dose_candidates = [_extract_dose(rec.dose), _extract_dose(rec.name), _extract_dose(row_blob)]
        dose_candidates = [d for d in dose_candidates if d]
        dose_candidate = ""
        if dose_candidates:
            complex_candidates = [d for d in dose_candidates if "+" in d]
            dose_candidate = max(complex_candidates or dose_candidates, key=len)
        if dose_candidate:
            rec.dose = dose_candidate

        if rec.price_per_unit and not _looks_like_price(rec.price_per_unit):
            rec.price_per_unit = ""
        if rec.sum_rub and not _looks_like_sum(rec.sum_rub):
            rec.sum_rub = ""

        compact = _extract_compact_values(row_blob)
        if not rec.qty_consumption_unit and compact.get("qty_consumption_unit"):
            rec.qty_consumption_unit = compact["qty_consumption_unit"]
        if not rec.price_per_unit and compact.get("price_per_unit"):
            rec.price_per_unit = compact["price_per_unit"]
        if not rec.sum_rub and compact.get("sum_rub"):
            rec.sum_rub = compact["sum_rub"]

        if rec.consumer_package_completeness and (
            rec.consumer_package_completeness.strip().startswith("~")
            or "СТРАНА ПРОИСХОЖДЕНИЯ" in rec.consumer_package_completeness.upper()
            or "ПРЕПАРАТ" in rec.consumer_package_completeness.upper()
        ):
            rec.consumer_package_completeness = "~"

        return rec

    def _merge_split_records(self, records: list[ParseRecord]) -> list[ParseRecord]:
        """Merge split records (e.g., main row + detail rows) but DO NOT merge duplicates.
        
        Each top-level table row is a separate position, even if names are identical.
        Only merge when a row is clearly an addon (metadata-only) for the previous core row.
        """
        if not records:
            return records

        merged: list[ParseRecord] = []
        last_core: Optional[ParseRecord] = None

        for rec in records:
            # If this is an addon row (metadata only), merge into the last core row
            if last_core and self._is_addon_row(rec):
                self._merge_into(last_core, rec)
                continue

            # Skip deduplication by key - each row from the HTML is a distinct position
            # This fixes the issue where identical drug names with different qty/price were merged
            merged.append(rec)
            if self._is_core_row(rec):
                last_core = rec

        return merged

    def _merge_into(self, base: ParseRecord, extra: ParseRecord) -> None:
        # Keep the most informative name while preventing OKPD2-only values.
        if extra.name and (not base.name or len(extra.name) > len(base.name) or _looks_like_okpd2_code(base.name)):
            if not _looks_like_okpd2_code(extra.name):
                base.name = extra.name

        for field in FIELD_ORDER:
            if field == "name":
                continue
            # Skip contract metadata fields - they are set at parse_url level
            if field in ('contract_date', 'contract_number', 'reestr_number', 'customer_name'):
                continue
            if not getattr(base, field) and getattr(extra, field):
                setattr(base, field, getattr(extra, field))

        # Price/sum can be mixed up in sparse rows. Keep stricter values if available in addon row.
        if extra.price_per_unit and _looks_like_price(extra.price_per_unit):
            base.price_per_unit = extra.price_per_unit
        if extra.sum_rub and _looks_like_sum(extra.sum_rub):
            base.sum_rub = _extract_sum(extra.sum_rub)

    def _record_key(self, rec: ParseRecord) -> str:
        key_id = _clean(rec.ru or rec.trade_name or rec.name)
        if not key_id:
            return ""
        return "|".join(
            [
                (_extract_okpd2(rec.okpd2) or _extract_okpd2(rec.name)).upper(),
                _short_country(rec.country).upper(),
                key_id.upper(),
            ]
        )

    def _is_core_row(self, rec: ParseRecord) -> bool:
        has_identity = bool(_clean(rec.trade_name) or _clean(rec.ru) or (_clean(rec.name) and not _looks_like_okpd2_code(rec.name)))
        has_top_cols = bool(_clean(rec.category_ls) or _clean(rec.okpd2) or _clean(rec.country))
        return has_identity or has_top_cols

    def _is_addon_row(self, rec: ParseRecord) -> bool:
        meta_fields = [
            rec.holder_name,
            rec.manufacturer_name,
            rec.manufacturer_country,
            rec.primary_package_type,
            rec.qty_forms_primary,
            rec.qty_primary_packages,
            rec.qty_consumer_units,
            rec.consumer_package_completeness,
        ]
        meta_count = sum(1 for v in meta_fields if _clean(v))
        weak_identity = not _clean(rec.trade_name) and not _looks_like_ru_code(rec.ru)
        return meta_count >= 2 and weak_identity

    def _split_compact_top_block(self, rec: ParseRecord, source_text: Optional[str] = None) -> None:
        src = source_text if source_text is not None else rec.country
        parts = [_clean(x) for x in str(src).split("|") if _clean(x)]
        if not parts:
            return

        if not rec.country:
            rec.country = _short_country(parts[0])
        for part in parts[1:]:
            up = part.upper()
            if "ТОВАР" == up:
                continue
            if "НДС" in up:
                if not rec.sum_rub:
                    rec.sum_rub = _extract_sum(part)
                continue
            if _extract_okpd2(part):
                if not rec.okpd2:
                    rec.okpd2 = _extract_okpd2(part)
                if not rec.category_ls:
                    rec.category_ls = _clean(re.sub(r"\(\d{2}\.\d{2}\.\d{2}\.\d{3}\)", "", part))
                continue
            if not rec.qty_consumption_unit and _looks_like_qty_text(part):
                rec.qty_consumption_unit = part
                continue
            if not rec.price_per_unit and _looks_like_price(part):
                rec.price_per_unit = _extract_price(part)
                continue
            if not rec.sum_rub and _looks_like_sum(part):
                rec.sum_rub = _extract_sum(part)


def _clean(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _extract_okpd2(text: str) -> str:
    m = re.search(r"(\d{2}\.\d{2}\.\d{2}\.\d{3})", text or "")
    return m.group(1) if m else ""


def _looks_like_okpd2_code(text: str) -> bool:
    return bool(re.fullmatch(r"\d{2}\.\d{2}\.\d{2}\.\d{3}", _clean(text)))


def _extract_dose(text: str) -> str:
    src = text or ""
    # Базовые единицы измерения
    unit = r"(?:МГ|МЛ|МКГ|МГК|МГ/МЛ|МГ/Г|МГ/КГ|МГ/ДОЗА|МЛ/ДОЗА|МКГ/ДОЗА|МГ/КВ\.СМ|МГ/СУТ|ЕД|МЕ|КОЕ|ЕМД|Г|%)"
    # Опциональный суффикс /<единица> для составных единиц измерения
    per_unit = r"(?:\s*/\s*(?:МЛ|Г|КГ|Л|МКГ|МГ|ЕД|МЕ|ДОЗА|КВ\.СМ|СУТ))?"
    # Полная единица измерения: МГ или МГ/МЛ или МЛ и т.д.
    full_unit = rf"{unit}{per_unit}"
    
    # Паттерн 1: Сложные дозировки с полной единицей в каждой части (10 МГ/МЛ+1 МГ/МЛ)
    complex_pat = re.compile(
        rf"((?:\d+[\.,]?\d*\s*{full_unit}\s*\+\s*)+\d+[\.,]?\d*\s*{full_unit})",
        flags=re.IGNORECASE,
    )
    complex_matches = [
        _clean(m.group(1))
        for m in complex_pat.finditer(src)
        if m and _clean(m.group(1))
    ]
    if complex_matches:
        return max(complex_matches, key=len)
    
    # Паттерн 2: Сложные дозировки с единицей только в конце (1500+500 мг, 0.72+5.75+2.7+0.36+0.05 мг/мл)
    complex_suffix_pat = re.compile(
        rf"((?:\d+[\.,]?\d*\s*\+\s*)+\d+[\.,]?\d*\s*{full_unit})",
        flags=re.IGNORECASE,
    )
    complex_suffix_matches = [
        _clean(m.group(1))
        for m in complex_suffix_pat.finditer(src)
        if m and _clean(m.group(1))
    ]
    if complex_suffix_matches:
        return max(complex_suffix_matches, key=len)
        
    # Фоллбеки для простых дозировок
    patterns = [
        rf"(\d+[\.,]?\d*\s*{full_unit})",
    ]
    for p in patterns:
        m = re.search(p, src, flags=re.IGNORECASE)
        if m:
            return _clean(m.group(1))
    return ""


def _extract_country(text: str) -> str:
    m = re.search(r"Страна происхождения\s*:\s*(.+)$", text or "", flags=re.IGNORECASE | re.MULTILINE)
    if not m:
        return ""
    # Извлекаем текст после метки и находим первую страну с кодом (XXX (NNN))
    country_text = _clean(m.group(1))
    country_match = re.search(r"([A-Za-zА-Яа-яЁё][A-Za-zА-Яа-яЁё\s]*\(\d{3}\))", country_text)
    if country_match:
        return _clean(country_match.group(1))
    return ""


def _short_country(text: str) -> str:
    src = _clean(text)
    if not src:
        return ""
    # Check for pattern with country code (XXX (NNN)) where XXX starts with a letter
    m = re.search(r"([A-Za-zА-Яа-яЁё][A-Za-zА-Яа-яЁё\s]*\(\d{3}\))", src)
    if m:
        return _clean(m.group(1))
    return ""


def _looks_like_price(text: str) -> bool:
    src = _clean(text)
    if not src or _looks_like_okpd2_code(src):
        return False
    if src.count(".") > 1:
        return False
    return bool(re.fullmatch(r"\d{1,6}[\.,]\d{2,7}", src))


def _looks_like_sum(text: str) -> bool:
    src = _clean(text)
    if not src:
        return False
    # Сумма может быть указана с НДС или без явного указания НДС в тексте
    # Проверяем наличие числа в формате суммы (с пробелами-разделителями тысяч и копейками)
    has_sum_pattern = bool(re.search(r"\d{1,3}(?:\s\d{3})+,\d{2}|\d{4,},\d{2}", src))
    has_vat = "НДС" in src.upper()
    # Возвращаем True если есть паттерн суммы и либо есть НДС, либо число похоже на итоговую сумму
    if has_vat:
        return has_sum_pattern
    # Если нет НДС в тексте, проверяем что это большое число (больше 1000)
    if has_sum_pattern:
        m = re.search(r"(\d{1,3}(?:\s\d{3})+,\d{2}|\d{4,},\d{2})", src)
        if m:
            num_str = m.group(1).replace(" ", "").replace(",", ".")
            try:
                if float(num_str) >= 1000:
                    return True
            except ValueError:
                pass
    return False


def _extract_price(text: str) -> str:
    src = _clean(text)
    for m in re.finditer(r"\b\d{1,6}[\.,]\d{2,7}\b", src):
        val = _clean(m.group(0))
        if _looks_like_okpd2_code(val):
            continue
        after = src[m.end() : m.end() + 12].upper()
        before = src[max(0, m.start() - 2) : m.start()].upper()
        # Skip dosage fragments like "0.400 мг+" or "3.2 мг/мл".
        if re.match(r"\s*(МГ|МЛ|МКГ|ЕД|МЕ|Г|КГ)\b", after):
            continue
        if "/" in after[:4] or "+" in before:
            continue
        # Ignore fragments from OKPD2-like chains such as 21.20.10.141
        if re.search(r"\d+\.\d+\.\d+\.\d+", src):
            if f".{val}" in src or f"{val}." in src:
                continue
        return val
    return ""


def _extract_sum(text: str) -> str:
    src = _clean(text)
    if "НДС" not in src.upper():
        return ""
    m = re.search(r"(\d{1,3}(?:\s\d{3})+,\d{2}|\d{4,},\d{2})", src)
    return _clean(m.group(1)) if m else ""


def _looks_like_qty_text(text: str) -> bool:
    src = _clean(text)
    if not src:
        return False
    if re.search(r"\b(ШТУКА\s*\(ШТ\)|ГРАММ\s*\(Г\)|МИЛЛИГРАММ\s*\(МГ\)|КУБИЧЕСКИЙ\s+САНТИМЕТР[^|]*\(СМ\[?3\*?\][^)]*\)|МИЛЛИЛИТР\s*\(МЛ\)|ШТ\.?|СМ\[?3\*?\])\b", src, flags=re.IGNORECASE):
        return True
    return bool(re.search(r"\d[\d\s.,]*\s*(СМ3|МЛ|Л|Г|КГ|ШТ|ЕД|МЕ)", src, flags=re.IGNORECASE))


def _looks_like_ru_code(text: str) -> bool:
    src = _clean(text)
    if not src:
        return False
    return bool(
        re.search(r"ЛП[-№A-ZА-Я0-9()\-/]+", src, flags=re.IGNORECASE)
        or re.search(r"\b[РP]?\s*№?\s*N?\d{5,}/\d+\b", src, flags=re.IGNORECASE)
    )


def _extract_compact_values(text: str) -> dict[str, str]:
    src = _clean(text)
    out = {"qty_consumption_unit": "", "price_per_unit": "", "sum_rub": ""}
    if not src:
        return out

    has_vat = "НДС" in src.upper()

    parts = [_clean(x) for x in src.split("|") if _clean(x)]
    for part in parts:
        up = part.upper()
        if not out["sum_rub"] and "НДС" in up:
            out["sum_rub"] = _extract_sum(part)
            continue
        if not out["qty_consumption_unit"] and _looks_like_qty_text(part):
            if len(part) > 100 and re.search(r"\d+\.\s+", part):
                continue
            out["qty_consumption_unit"] = part
            continue
        if not out["price_per_unit"] and _looks_like_price(part):
            out["price_per_unit"] = _extract_price(part)
            continue
        if has_vat and not out["sum_rub"] and _looks_like_sum(part):
            out["sum_rub"] = _extract_sum(part)

    if not out["qty_consumption_unit"]:
        out["qty_consumption_unit"] = _extract_qty_from_blob(src)
    if out["qty_consumption_unit"] and len(out["qty_consumption_unit"]) > 120:
        # Some pages leak full item description into qty cell; keep only unit text.
        trimmed_qty = _extract_qty_from_blob(out["qty_consumption_unit"])
        out["qty_consumption_unit"] = trimmed_qty if trimmed_qty and len(trimmed_qty) < len(out["qty_consumption_unit"]) else ""

    # Global fallback for rows where compact values are merged without "|" delimiters.
    if has_vat and not out["sum_rub"]:
        sum_match = re.search(r"\b(\d{1,3}(?:\s\d{3})+,\d{2}|\d{4,},\d{2})\b", src)
        if sum_match:
            out["sum_rub"] = _clean(sum_match.group(1))

    if not out["price_per_unit"]:
        for m in re.finditer(r"\b\d{1,6}[\.,]\d{2,7}\b", src):
            val = _clean(m.group(0))
            if _looks_like_okpd2_code(val):
                continue
            after = src[m.end() : m.end() + 12].upper()
            if re.match(r"\s*(МГ|МЛ|МКГ|ЕД|МЕ|Г|КГ)\b", after):
                continue
            if "/" in after[:4]:
                continue
            if out["sum_rub"] and val == out["sum_rub"]:
                continue
            if re.search(r"\d+\.\d+\.\d+\.\d+", src) and (f".{val}" in src or f"{val}." in src):
                continue
            out["price_per_unit"] = val
            break

    # Last resort: infer from decimal sequence when source block is compact without separators.
    if not out["price_per_unit"] or (has_vat and not out["sum_rub"]):
        nums: list[str] = []
        for m in re.finditer(r"\b\d{1,6}[\.,]\d{2,7}\b", src):
            val = _clean(m.group(0))
            if _looks_like_okpd2_code(val):
                continue
            after = src[m.end() : m.end() + 12].upper()
            if re.match(r"\s*(МГ|МЛ|МКГ|ЕД|МЕ|Г|КГ)\b", after):
                continue
            if "/" in after[:4]:
                continue
            nums.append(val)
        if nums:
            if not out["price_per_unit"]:
                out["price_per_unit"] = nums[0]
            if has_vat and not out["sum_rub"] and len(nums) > 1:
                out["sum_rub"] = nums[-1]

    if not out["price_per_unit"]:
        out["price_per_unit"] = _extract_price(src)
    if has_vat and not out["sum_rub"]:
        out["sum_rub"] = _extract_sum(src)
    return out


def _extract_name_from_blob(text: str) -> str:
    src = _clean(text)
    if not src:
        return ""

    m = re.search(r"(\d+\.\s*[^|]+)", src, flags=re.IGNORECASE)
    if not m:
        return ""

    value = _clean(m.group(1))
    stop_markers = [
        "ЕДИНИЦА ИЗМЕРЕНИЯ ТОВАРА",
        "СТРАНА ПРОИСХОЖДЕНИЯ",
        "МНН И ФОРМА ВЫПУСКА",
        "ТОРГОВОЕ НАИМЕНОВАНИЕ",
    ]
    up = value.upper()
    for marker in stop_markers:
        idx = up.find(marker)
        if idx > 0:
            value = _clean(value[:idx])
            up = value.upper()
    return value


def _extract_qty_from_blob(text: str) -> str:
    src = _clean(text)
    if not src:
        return ""

    labeled = re.search(r"ЕДИНИЦА\s+ИЗМЕРЕНИЯ(?:\s+ТОВАРА)?\s*:\s*([^|]+)", src, flags=re.IGNORECASE)
    if labeled:
        value = _clean(labeled.group(1))
        value = re.split(
            r"СТРАНА\s+ПРОИСХОЖДЕНИЯ|ПРЕПАРАТЫ|ТОВАР\b|НАИМЕНОВАНИЕ\s+ДЕРЖАТЕЛЯ|НАИМЕНОВАНИЕ\s+ПРОИЗВОДИТЕЛЯ",
            value,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0]
        return _clean(value)

    if "|" in src:
        for part in [_clean(x) for x in src.split("|") if _clean(x)]:
            if _looks_like_qty_text(part):
                return part

    unit_only = re.search(
        r"(ШТУКА\s*\(ШТ\)|КУБИЧЕСКИЙ\s+САНТИМЕТР[^|]*\(СМ\[?3\*?\][^)]*\)|ГРАММ\s*\(Г\)|МИЛЛИГРАММ\s*\(МГ\)|МИЛЛИЛИТР\s*\(МЛ\))",
        src,
        flags=re.IGNORECASE,
    )
    if unit_only:
        return _clean(unit_only.group(1))

    m = re.search(r"(\d[\d\s.,]*\s*(?:СМ3|МЛ|Л|Г|КГ|ШТ|ЕД|МЕ)[^|]*)", src, flags=re.IGNORECASE)
    return _clean(m.group(1)) if m else ""


def _looks_like_meta_blob(text: str) -> bool:
    src = _clean(text).upper()
    keys = [
        "НАИМЕНОВАНИЕ ДЕРЖАТЕЛЯ ИЛИ ВЛАДЕЛЬЦА РУ",
        "НАИМЕНОВАНИЕ ПРОИЗВОДИТЕЛЯ",
        "ВИД ПЕРВИЧНОЙ УПАКОВКИ",
        "КОМПЛЕКТНОСТЬ ПОТРЕБИТЕЛЬСКОЙ УПАКОВКИ",
    ]
    return any(k in src for k in keys)


def _extract_meta_fields(text: str) -> dict[str, str]:
    src = _clean(text).replace("|", " ")
    labels = {
        "release_form": "Лекарственная форма",
        "dose": "Дозировка",
        "holder_name": "Наименование держателя или владельца РУ",
        "manufacturer_name": "Наименование производителя",
        "manufacturer_country": "Страна производителя",
        "primary_package_type": "Вид первичной упаковки",
        "qty_forms_primary": "Количество лекарственных форм в первичной упаковке",
        "qty_primary_packages": "Количество первичных упаковок в потребительской упаковке",
        "qty_consumer_units": "Количество потребительских единиц в потребительской упаковке",
        "consumer_package_completeness": "Комплектность потребительской упаковки",
    }
    all_markers = sorted((re.escape(v) for v in labels.values()), key=len, reverse=True)
    any_marker = "|".join(all_markers)

    out: dict[str, str] = {}
    for field, label in labels.items():
        pattern = re.compile(
            rf"{re.escape(label)}\s*:?\s*(.+?)(?=(?:{any_marker})\s*:?|$)",
            flags=re.IGNORECASE,
        )
        m = pattern.search(src)
        if m:
            out[field] = _clean(m.group(1))
    return out


def _read_urls(single_url: Optional[str], url_file: Optional[Path]) -> list[str]:
    urls: list[str] = []
    if single_url:
        urls.append(single_url.strip())
    if url_file and url_file.exists():
        for line in url_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                urls.append(line)
    if not urls:
        raise ValueError("Передайте --url или --url-file")
    return urls


def export_csv(rows: list[dict[str, str]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELD_ORDER)
        writer.writeheader()
        writer.writerows(rows)


def export_xlsx(rows: list[dict[str, str]], out_path: Path) -> bool:
    try:
        import pandas as pd
        from openpyxl import load_workbook
    except Exception:
        return False

    out_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows, columns=FIELD_ORDER)
    df = df.rename(columns=EXPORT_HEADERS_RU)
    df.to_excel(out_path, index=False)
    
    # Применяем числовой формат к указанным полям
    numeric_columns = [
        "Цена за единицу измерения, ₽",
        "Количество лекарственных форм в первичной упаковке",
        "Количество первичных упаковок в потребительской упаковке",
        "Количество потребительских единиц в потребительской упаковке"
    ]
    
    wb = load_workbook(out_path)
    ws = wb.active
    
    # Находим индексы столбцов для числовых полей
    header_row = 1
    col_indices = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value in numeric_columns:
            col_indices[cell.value] = col_idx
    
    # Применяем числовой формат к данным
    for row_idx in range(2, ws.max_row + 1):
        for col_name, col_idx in col_indices.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and cell.value != "":
                try:
                    # Преобразуем значение в число
                    numeric_value = float(str(cell.value).replace(" ", "").replace(",", "."))
                    cell.value = numeric_value
                    # Применяем числовой формат
                    cell.number_format = '0.00' if "Цена" in col_name else '0'
                except (ValueError, TypeError):
                    pass
    
    wb.save(out_path)
    return True


def _log_line(log_fn, text: str, *, is_error: bool = False) -> None:
    if log_fn:
        log_fn(text, is_error=is_error)
    else:
        stream = sys.stderr if is_error else sys.stdout
        print(text, file=stream)


async def run(args: argparse.Namespace, log_fn=None) -> int:
    parser = EISParser(
        timeout_ms=args.timeout_ms,
        expand_rounds=args.expand_rounds,
        page_load_delay=getattr(args, 'page_load_delay', 1200),
        expand_delay=getattr(args, 'expand_delay', 800),
    )
    archive_dir = Path(args.archive_dir)
    urls = _read_urls(args.url, Path(args.url_file) if args.url_file else None)

    async with async_playwright() as p:
        browser: Browser = await p.chromium.launch(headless=not args.headed)
        context: BrowserContext = await browser.new_context(locale="ru-RU")
        if args.trace:
            await context.tracing.start(screenshots=True, snapshots=True, sources=True)

        all_rows: list[dict[str, str]] = []
        for idx, url in enumerate(urls, start=1):
            page = await context.new_page()
            try:
                _log_line(log_fn, f"[{idx}/{len(urls)}] Parse: {url}")
                rows = await parser.parse_url(page, url, archive_dir=archive_dir, save_trace=args.trace)
                all_rows.extend(rows)
                _log_line(log_fn, f"  -> rows: {len(rows)}")
            except Exception as exc:
                fail_dir = archive_dir / "failures"
                fail_dir.mkdir(parents=True, exist_ok=True)
                safe_name = re.sub(r"[^\w\-]+", "_", url)[:120]
                await page.screenshot(path=str(fail_dir / f"{safe_name}.png"), full_page=True)
                (fail_dir / f"{safe_name}.txt").write_text(str(exc), encoding="utf-8")
                _log_line(log_fn, f"  -> error: {exc}", is_error=True)
            finally:
                await page.close()

        if args.trace:
            trace_name = archive_dir / f"trace_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            await context.tracing.stop(path=str(trace_name))
            _log_line(log_fn, f"Trace: {trace_name}")

        await context.close()
        await browser.close()

    if not all_rows:
        _log_line(log_fn, "Нет данных для экспорта", is_error=True)
        return 2

    csv_out = Path(args.out_csv)
    export_csv(all_rows, csv_out)
    _log_line(log_fn, f"CSV saved: {csv_out}")

    if args.out_xlsx:
        ok = export_xlsx(all_rows, Path(args.out_xlsx))
        if ok:
            _log_line(log_fn, f"XLSX saved: {args.out_xlsx}")
        else:
            _log_line(log_fn, "XLSX skipped: установите pandas и openpyxl", is_error=True)

    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="EIS parser: Объекты закупки")
    p.add_argument("--gui", action="store_true", help="Запуск в графическом интерфейсе")
    p.add_argument("--url", help="Single EIS URL")
    p.add_argument("--url-file", help="Path to file with one URL per line")
    p.add_argument("--archive-dir", default="archive", help="Debug archive directory")
    p.add_argument("--out-csv", default="export/result.csv", help="Output CSV path")
    p.add_argument("--out-xlsx", default="export/result.xlsx", help="Output XLSX path")
    p.add_argument("--headed", action="store_true", help="Run browser in headed mode")
    p.add_argument("--trace", action="store_true", help="Save Playwright trace.zip")
    p.add_argument("--timeout-ms", type=int, default=60000, help="Navigation timeout")
    p.add_argument("--expand-rounds", type=int, default=4, help="Expand passes for nested rows")
    p.add_argument("--page-load-delay", type=int, default=1200, help="Delay after page load (ms)")
    p.add_argument("--expand-delay", type=int, default=800, help="Delay after expand operations (ms)")
    return p


class EISParserGUI(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("EIS Parser - Единый интерфейс")
        self.geometry("1380x800")
        self.minsize(1120, 640)

        self.url_var = tk.StringVar(
            value=(
                "https://zakupki.gov.ru/epz/contract/contractCard/payment-info-and-target-of-order.html"
                "?reestrNumber=2312813818126000251&contractInfoId=108730614"
            )
        )
        self.search_var = tk.StringVar(value="")
        self.archive_var = tk.StringVar(value="archive")
        self.csv_var = tk.StringVar(value="export/result.csv")
        self.xlsx_var = tk.StringVar(value="export/result.xlsx")
        self.trace_var = tk.BooleanVar(value=True)
        self.headed_var = tk.BooleanVar(value=False)
        self.timeout_ms_var = tk.IntVar(value=90000)
        self.expand_rounds_var = tk.IntVar(value=5)
        self.page_load_delay_var = tk.IntVar(value=1200)
        self.expand_delay_var = tk.IntVar(value=800)

        self.all_rows: list[dict[str, str]] = []
        self.filtered_rows: list[dict[str, str]] = []
        self._worker_thread: Optional[threading.Thread] = None

        self._build_ui()

    def _build_ui(self) -> None:
        root = ttk.Frame(self, padding=10)
        root.pack(fill=tk.BOTH, expand=True)

        title = ttk.Label(
            root,
            text="Запуск парсинга ЕИС в одном окне",
            font=("Segoe UI", 13, "bold"),
        )
        title.pack(anchor="w", pady=(0, 8))

        controls = ttk.LabelFrame(root, text="Параметры", padding=8)
        controls.pack(fill=tk.X)

        self._row_input(controls, "Ссылка ЕИС", self.url_var, add_paste_button=True)
        self._row_input(controls, "Слово для фильтра (опц.)", self.search_var, on_change=self._apply_filter)
        self._row_input(controls, "Папка архива", self.archive_var)
        self._row_input(controls, "CSV файл", self.csv_var)
        self._row_input(controls, "XLSX файл", self.xlsx_var)

        # Timing settings frame
        timing_frame = ttk.LabelFrame(controls, text="Тайминги и скорость обработки", padding=6)
        timing_frame.pack(fill=tk.X, pady=(6, 0))
        
        self._row_spinbox(timing_frame, "Таймаут загрузки (мс):", self.timeout_ms_var, from_=30000, to=300000, step=10000)
        self._row_spinbox(timing_frame, "Раунды раскрытия:", self.expand_rounds_var, from_=1, to=10, step=1)
        self._row_spinbox(timing_frame, "Задержка после загрузки (мс):", self.page_load_delay_var, from_=200, to=5000, step=100)
        self._row_spinbox(timing_frame, "Задержка раскрытия (мс):", self.expand_delay_var, from_=200, to=3000, step=100)

        toggles = ttk.Frame(controls)
        toggles.pack(fill=tk.X, pady=(4, 0))
        ttk.Checkbutton(toggles, text="Сохранять trace", variable=self.trace_var).pack(side=tk.LEFT)
        ttk.Checkbutton(
            toggles,
            text="Показывать окно браузера (headed)",
            variable=self.headed_var,
        ).pack(side=tk.LEFT, padx=(14, 0))

        actions = ttk.Frame(root)
        actions.pack(fill=tk.X, pady=(8, 8))
        self.run_btn = ttk.Button(actions, text="Запустить парсинг", command=self._start_parse)
        self.run_btn.pack(side=tk.LEFT)
        ttk.Button(actions, text="Открыть CSV", command=self._open_csv_file).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(actions, text="Выгрузить Excel", command=self._export_excel_file).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(actions, text="Очистить таблицу", command=self._clear_rows).pack(side=tk.LEFT, padx=(8, 0))

        self.status_var = tk.StringVar(value="Готово")
        ttk.Label(root, textvariable=self.status_var, foreground="#1f4e79").pack(anchor="w", pady=(0, 6))

        content = ttk.Panedwindow(root, orient=tk.VERTICAL)
        content.pack(fill=tk.BOTH, expand=True)

        top = ttk.Frame(content)
        bottom = ttk.Frame(content)
        content.add(top, weight=4)
        content.add(bottom, weight=2)

        columns = tuple(FIELD_ORDER)
        self.tree = ttk.Treeview(top, columns=columns, show="headings", height=18)
        for field in FIELD_ORDER:
            self.tree.heading(field, text=EXPORT_HEADERS_RU[field])
            self.tree.column(field, width=150 if field != "name" else 270, anchor="w")

        sy = ttk.Scrollbar(top, orient=tk.VERTICAL, command=self.tree.yview)
        sx = ttk.Scrollbar(top, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sy.pack(side=tk.RIGHT, fill=tk.Y)
        sx.pack(side=tk.BOTTOM, fill=tk.X)

        ttk.Label(bottom, text="Лог выполнения").pack(anchor="w")
        self.log_text = tk.Text(bottom, height=9, wrap="word")
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def _row_input(self, parent, label, variable, on_change=None, add_paste_button: bool = False) -> None:
        row = ttk.Frame(parent)
        row.pack(fill=tk.X, pady=2)
        ttk.Label(row, text=label, width=22).pack(side=tk.LEFT)
        ent = ttk.Entry(row, textvariable=variable)
        ent.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._add_context_menu(ent)
        ent.bind("<FocusIn>", lambda _e: ent.selection_range(0, tk.END))
        if add_paste_button:
            ttk.Button(
                row,
                text="Вставить",
                command=lambda: self._paste_from_clipboard(variable),
            ).pack(side=tk.LEFT, padx=(6, 0))
        if on_change:
            ent.bind("<KeyRelease>", lambda _e: on_change())

    def _row_spinbox(self, parent, label, variable, from_=0, to=100, step=1) -> None:
        row = ttk.Frame(parent)
        row.pack(fill=tk.X, pady=2)
        ttk.Label(row, text=label, width=28).pack(side=tk.LEFT)
        spin = ttk.Spinbox(row, textvariable=variable, from_=from_, to=to, increment=step, width=15)
        spin.pack(side=tk.LEFT, fill=tk.X, expand=True)
        spin.bind("<FocusIn>", lambda _e: spin.selection_range(0, tk.END))

    def _add_context_menu(self, widget: ttk.Entry) -> None:
        menu = tk.Menu(widget, tearoff=0)
        menu.add_command(label="Копировать", command=lambda: widget.event_generate("<<Copy>>"))
        menu.add_command(label="Вырезать", command=lambda: widget.event_generate("<<Cut>>"))
        menu.add_command(label="Вставить", command=lambda: widget.event_generate("<<Paste>>"))
        menu.add_separator()
        menu.add_command(label="Выделить все", command=lambda: widget.selection_range(0, tk.END))

        def show_menu(event):
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()

        widget.bind("<Button-3>", show_menu)
        widget.bind("<Control-Button-1>", show_menu)

    def _paste_from_clipboard(self, variable: tk.StringVar) -> None:
        try:
            text = self.clipboard_get().strip()
        except tk.TclError:
            messagebox.showwarning("Внимание", "Буфер обмена пуст или недоступен")
            return
        variable.set(text)

    def _set_status(self, text: str) -> None:
        self.status_var.set(text)
        self.update_idletasks()

    def _append_log(self, text: str, is_error: bool = False) -> None:
        prefix = "[ERR] " if is_error else ""
        self.log_text.insert("end", f"{prefix}{text}\n")
        self.log_text.see("end")

    def _thread_log(self, text: str, is_error: bool = False) -> None:
        self.after(0, lambda: self._append_log(text, is_error=is_error))

    def _start_parse(self) -> None:
        if self._worker_thread and self._worker_thread.is_alive():
            messagebox.showinfo("Информация", "Парсинг уже выполняется")
            return

        url = self.url_var.get().strip()
        if not url:
            messagebox.showwarning("Внимание", "Укажите ссылку ЕИС")
            return

        self.run_btn.config(state=tk.DISABLED)
        self._set_status("Запуск парсинга...")
        self._append_log("Запуск...")

        config = {
            "url": url,
            "archive_dir": self.archive_var.get().strip() or "archive",
            "out_csv": self.csv_var.get().strip() or "export/result.csv",
            "out_xlsx": self.xlsx_var.get().strip() or "",
            "headed": self.headed_var.get(),
            "trace": self.trace_var.get(),
        }

        self._worker_thread = threading.Thread(target=self._worker, args=(config,), daemon=True)
        self._worker_thread.start()

    def _worker(self, config: dict) -> None:
        ns = argparse.Namespace(
            gui=False,
            url=config["url"],
            url_file=None,
            archive_dir=config["archive_dir"],
            out_csv=config["out_csv"],
            out_xlsx=config["out_xlsx"],
            headed=config["headed"],
            trace=config["trace"],
            timeout_ms=self.timeout_ms_var.get(),
            expand_rounds=self.expand_rounds_var.get(),
            page_load_delay=self.page_load_delay_var.get(),
            expand_delay=self.expand_delay_var.get(),
        )
        try:
            rc = asyncio.run(run(ns, log_fn=self._thread_log))
        except Exception as exc:
            self.after(0, lambda: self._append_log(f"Критическая ошибка: {exc}", is_error=True))
            rc = 1

        self.after(0, lambda: self._on_worker_done(rc, Path(ns.out_csv)))

    def _on_worker_done(self, rc: int, csv_path: Path) -> None:
        self.run_btn.config(state=tk.NORMAL)
        if rc == 0 and csv_path.exists():
            self._load_rows_from_csv(csv_path)
            # Always show parse result first. Search is only optional client-side filtering.
            self.filtered_rows = list(self.all_rows)
            self._render_rows()
            self._set_status(f"Готово. Строк: {len(self.all_rows)}")
        else:
            self._set_status("Завершено с ошибкой")

    def _load_rows_from_csv(self, csv_path: Path) -> None:
        rows: list[dict[str, str]] = []
        with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for item in reader:
                rows.append({k: _clean(item.get(k, "")) for k in FIELD_ORDER})
        self.all_rows = rows

    def _apply_filter(self) -> None:
        query = self.search_var.get().strip().lower()
        if not query:
            self.filtered_rows = list(self.all_rows)
        else:
            self.filtered_rows = [
                r for r in self.all_rows if any(query in str(v).lower() for v in r.values())
            ]
        self._render_rows()

    def _render_rows(self) -> None:
        for item in self.tree.get_children():
            self.tree.delete(item)

        for r in self.filtered_rows:
            self.tree.insert("", "end", values=[r.get(field, "") for field in FIELD_ORDER])

    def _clear_rows(self) -> None:
        self.all_rows = []
        self.filtered_rows = []
        self._render_rows()
        self._set_status("Очищено")

    def _open_csv_file(self) -> None:
        path = Path(self.csv_var.get().strip() or "export/result.csv")
        if not path.exists():
            messagebox.showwarning("Внимание", "CSV файл еще не создан")
            return
        try:
            import os

            os.startfile(path.resolve())  # type: ignore[attr-defined]
        except Exception as exc:
            messagebox.showerror("Ошибка", f"Не удалось открыть CSV: {exc}")

    def _export_excel_file(self) -> None:
        rows = self.filtered_rows if self.filtered_rows else self.all_rows
        if not rows:
            messagebox.showwarning("Внимание", "Нет данных для выгрузки в Excel")
            return

        out_path = Path(self.xlsx_var.get().strip() or "export/result.xlsx")
        ok = export_xlsx(rows, out_path)
        if not ok:
            messagebox.showerror(
                "Ошибка",
                "Не удалось выгрузить Excel. Установите pandas и openpyxl:\n"
                "pip install pandas openpyxl",
            )
            return

        self._append_log(f"XLSX saved: {out_path}")
        self._set_status(f"Excel выгружен: {out_path}")
        messagebox.showinfo("Готово", f"Excel файл сохранен:\n{out_path}")


def launch_gui() -> int:
    app = EISParserGUI()
    app.mainloop()
    return 0


def main() -> int:
    args = build_arg_parser().parse_args()
    if args.gui:
        return launch_gui()
    try:
        return asyncio.run(run(args))
    except KeyboardInterrupt:
        print("Interrupted by user", file=sys.stderr)
        return 130


EXTRACT_SCRIPT = r"""
() => {
  const clean = (s) => (s || '').replace(/\u00A0/g, ' ').replace(/\s+/g, ' ').trim();
  const text = (node) => clean((node?.innerText || node?.textContent || ''));

  const rx = {
    okpd2: /(\d{2}\.\d{2}\.\d{2}\.\d{3})/,
    ru: /(ЛП[-№\w()\/]+|\b[РP]?\s*№?\s*N?\d{5,}\/\d+\b)/i,
    dose: /((?:\d+[\.,]?\d*\s*(?:МГ|МЛ|МКГ|ЕД|МЕ|Г)\s*\+\s*)+\d+[\.,]?\d*\s*(?:МГ|МЛ|МКГ|ЕД|МЕ|Г)(?:\s*\/\s*(?:МЛ|Г|КГ|Л|МКГ|МГ|ЕД|МЕ))?|\d+[\.,]?\d*\s*(?:МГ|МЛ|МКГ|ЕД|МЕ|Г)\s*\/\s*(?:МЛ|Г|КГ|Л|МКГ|МГ|ЕД|МЕ)|\d+[\.,]?\d*\s*(?:МГ|МЛ|МКГ|ЕД|МЕ|Г))/i,
    countryShort: /([A-Za-zА-Яа-яЁё\-\s]+\(\d{3}\))/,
    sum: /(\d{1,3}(?:\s\d{3})+,\d{2}|\d+,\d{2})/
  };

  const blank = () => ({
    name: '',
    category_ls: '',
    okpd2: '',
    country: '',
    mnn: '',
    trade_name: '',
    ru: '',
    release_form: '',
    dose: '',
    qty_consumption_unit: '',
    price_per_unit: '',
    sum_rub: '',
    nds: '',
    holder_name: '',
    manufacturer_name: '',
    manufacturer_country: '',
    primary_package_type: '',
    qty_forms_primary: '',
    qty_primary_packages: '',
    qty_consumer_units: '',
    consumer_package_completeness: '',
  });

  const looksLikePrice = (s) => /^\d{1,6}[.,]\d{2,7}$/.test(clean(s));
  const looksLikeSumRub = (s) => /^\d{1,3}(\s\d{3})+,\d{2}$/.test(clean(s)) || /^\d{4,},\d{2}$/.test(clean(s));
  const looksLikeQty = (s) => {
    const src = clean(s);
    return /\d[\d\s.,]*\s*(СМ3|МЛ|Л|Г|КГ|ШТ|ЕД|МЕ)/i.test(src)
      || /(Штука\s*\(шт\)|Кубический\s+сантиметр[^|]*\(см\[?3\*?\][^)]*\)|Миллилитр\s*\(мл\)|Грамм\s*\(г\)|Миллиграмм\s*\(мг\))/i.test(src);
  };

  const shortCountry = (s) => {
    const src = clean(s);
    if (!src) return '';
    // Check for pattern with country code (XXX (NNN)) where XXX starts with a letter
    const m = src.match(rx.countryShort);
    if (m) {
      return clean(m[1]);
    }
    return '';
  };

  const extractSum = (s) => clean((clean(s).match(rx.sum) || [])[1] || '');

  const extractPrice = (srcText) => {
    const src = clean(srcText);
    if (!src) return '';
    const candidates = src.match(/\d{1,6}[.,]\d{2,7}/g) || [];
    for (const c of candidates) {
      const val = clean(c);
      if (!val) continue;
      if (/^\d{2}\.\d{2}\.\d{2}\.\d{3}$/.test(val)) continue;
      const idx = src.indexOf(val);
      const after = idx >= 0 ? src.slice(idx + val.length, idx + val.length + 12).toUpperCase() : '';
      if (/^\s*(МГ|МЛ|МКГ|ЕД|МЕ|Г|КГ)\b/.test(after)) continue;
      if (after.slice(0, 4).includes('/')) continue;
      if (/\d+\.\d+\.\d+\.\d+/.test(src) && (src.includes(`.${val}`) || src.includes(`${val}.`))) {
        continue;
      }
      return val;
    }
    return '';
  };

  const extractQty = (srcText) => {
    const src = clean(srcText);
    if (!src) return '';
    const labeled = src.match(/ЕДИНИЦА\s+ИЗМЕРЕНИЯ(?:\s+ТОВАРА)?\s*:\s*([^|]+)/i);
    if (labeled) {
      return clean(labeled[1].split(/СТРАНА\s+ПРОИСХОЖДЕНИЯ/i)[0]);
    }
    const unitOnly = src.match(/(Штука\s*\(шт\)|Кубический\s+сантиметр[^|]*\(см\[?3\*?\][^)]*\)|Миллилитр\s*\(мл\)|Грамм\s*\(г\)|Миллиграмм\s*\(мг\))/i);
    if (unitOnly) return clean(unitOnly[1]);
    const m = src.match(/\d[\d\s.,]*\s*(СМ3|МЛ|Л|Г|КГ|ШТ|ЕД|МЕ)(?:\s*\([^)]*\))?/i);
    return m ? clean(m[0]) : '';
  };

  const extractName = (srcText) => {
    const src = clean(srcText);
    if (!src) return '';
    const m = src.match(/\d+\.\s*[^|]+/);
    if (!m) return '';
    let value = clean(m[0]);
    value = clean(value.replace(/Страна происхождения\s*:.*/i, ''));
    return value;
  };

  const metaMap = [
    ['release_form', 'Лекарственная форма'],
    ['dose', 'Дозировка'],
    ['holder_name', 'Наименование держателя или владельца РУ'],
    ['manufacturer_name', 'Наименование производителя'],
    ['manufacturer_country', 'Страна производителя'],
    ['primary_package_type', 'Вид первичной упаковки'],
    ['qty_forms_primary', 'Количество лекарственных форм в первичной упаковке'],
    ['qty_primary_packages', 'Количество первичных упаковок в потребительской упаковке'],
    ['qty_consumer_units', 'Количество потребительских единиц в потребительской упаковке'],
    ['consumer_package_completeness', 'Комплектность потребительской упаковки'],
  ];

  const extractMetaFromText = (srcText) => {
    const src = clean(srcText).replace(/\|/g, ' ');
    const points = [];
    for (const [key, label] of metaMap) {
      const token = `${label}:`;
      const idx = src.toUpperCase().indexOf(token.toUpperCase());
      if (idx >= 0) points.push({ key, label, idx, end: idx + token.length });
    }
    points.sort((a, b) => a.idx - b.idx);
    const out = {};
    for (let i = 0; i < points.length; i++) {
      const cur = points[i];
      const nextIdx = i + 1 < points.length ? points[i + 1].idx : src.length;
      out[cur.key] = clean(src.slice(cur.end, nextIdx));
    }
    return out;
  };

  const section = Array.from(document.querySelectorAll('section,div,table')).find((el) =>
    /Объекты закупки/i.test(text(el))
  ) || document.body;

  const isTopRow = (tr) => {
    // Check if row starts with a number pattern like "1." or "2."
    if (!/^\s*\d+\.\s+/.test(text(tr))) return false;
    const tds = tr.querySelectorAll('td');
    // Top rows have 7+ cells OR contain qty/price/sum columns
    if (tds.length >= 7) return true;
    // Also check if it contains quantity-like data in expected column positions
    const cellTexts = Array.from(tds).map(td => text(td));
    if (cellTexts.length >= 5) {
      // Check if any cell looks like quantity with unit
      for (const c of cellTexts) {
        if (looksLikeQty(c)) return true;
      }
    }
    return false;
  };
  const topRows = Array.from(section.querySelectorAll('tr')).filter((tr) => {
    return isTopRow(tr);
  });

  const extractMnnFromExpandedRow = (tr) => {
    // Look for MNN in expanded sibling row (ktru-tr-XXXXX pattern)
    let nextRow = tr.nextElementSibling;
    while (nextRow && nextRow.tagName === 'TR') {
      const rowId = nextRow.id || '';
      if (rowId.startsWith('ktru-tr-')) {
        // This is an expanded detail row
        const sections = nextRow.querySelectorAll('.blockInfo__section');
        for (const sec of sections) {
          const titleEl = sec.querySelector('.section__title');
          const infoEl = sec.querySelector('.section__info');
          if (!titleEl || !infoEl) continue;
          const title = clean(text(titleEl));
          const info = clean(text(infoEl));
          if (!info) continue;
          
          const titleUp = title.toUpperCase();
          if (/МЕЖДУНАРОДНОЕ.*НАИМЕНОВАНИЕ/.test(titleUp)) {
            return { mnn: info, row: nextRow };
          }
        }
        break;
      }
      nextRow = nextRow.nextElementSibling;
    }
    return { mnn: '', row: null };
  };
  
  const extractManufacturerCountryFromExpandedRow = (expandedRow) => {
    if (!expandedRow) return '';
    const sections = expandedRow.querySelectorAll('.blockInfo__section');
    for (const sec of sections) {
      const titleEl = sec.querySelector('.section__title');
      const infoEl = sec.querySelector('.section__info');
      if (!titleEl || !infoEl) continue;
      const title = clean(text(titleEl));
      const info = clean(text(infoEl));
      if (!info) continue;
      
      const titleUp = title.toUpperCase();
      if (titleUp.includes('СТРАНА ПРОИЗВОДИТЕЛЯ')) {
        return info;
      }
    }
    return '';
  };

  const parseTopRow = (tr) => {
    const rec = blank();
    const cols = Array.from(tr.querySelectorAll('td')).map((td) => text(td));
    const rowText = text(tr);
    rec.name = cols.map(extractName).find(Boolean) || extractName(rowText) || clean(cols[0] || '');

    // Extract MNN from expanded sibling row
    const { mnn, row: expandedRow } = extractMnnFromExpandedRow(tr);
    if (mnn) {
      rec.mnn = mnn;
    }
    
    // Extract manufacturer country from expanded sibling row
    const manufacturerCountry = extractManufacturerCountryFromExpandedRow(expandedRow);
    if (manufacturerCountry) {
      rec.manufacturer_country = manufacturerCountry;
    }

    // Direct column mapping for quantity, price, and sum when we have enough columns
    // Typical structure: [empty, name+country, category+okpd2, "Товар", qty, price, sum+VAT]
    // OR for rows with 9 cells: [empty, name, empty, trade, ru, form, dose, qty, meta]
    if (cols.length >= 7) {
      // Try standard 7-column layout first
      let foundQty = false;
      
      // Column 4 (index 4) typically contains quantity with unit
      const qtyCell = clean(cols[4] || '');
      if (qtyCell && looksLikeQty(qtyCell)) {
        rec.qty_consumption_unit = extractQty(qtyCell) || qtyCell;
        foundQty = true;
      }
      
      // Column 5 (index 5) typically contains price per unit
      const priceCell = clean(cols[5] || '');
      if (priceCell && looksLikePrice(priceCell)) {
        rec.price_per_unit = priceCell;
      }
      
      // Column 6 (index 6) typically contains sum with VAT info
      const sumCell = clean(cols[6] || '');
      if (sumCell && /НДС/i.test(sumCell)) {
        // Extract sum - first number with comma and 2 decimals before НДС
        const sumMatch = sumCell.match(/(\d{1,3}(?:\s\d{3})+,\d{2}|\d+,\d{2})/);
        if (sumMatch) {
          rec.sum_rub = clean(sumMatch[1]);
        }
        // Extract VAT rate from the cell - look for percentage after НДС
        const ndsMatch = sumCell.match(/НДС[:\s]*(\d+\s*%?)/i);
        if (ndsMatch) {
          let ndsVal = clean(ndsMatch[1]);
          if (!ndsVal.includes('%')) ndsVal = ndsVal + '%';
          rec.nds = ndsVal;
        }
      }
      
      // For 9-column layouts, check alternative positions
      if (!foundQty && cols.length >= 9) {
        // In some layouts, qty might be in a different position
        for (let i = 4; i < cols.length; i++) {
          const cell = clean(cols[i] || '');
          if (!rec.qty_consumption_unit && looksLikeQty(cell)) {
            rec.qty_consumption_unit = extractQty(cell) || cell;
          }
          if (!rec.price_per_unit && looksLikePrice(cell)) {
            rec.price_per_unit = cell;
          }
        }
      }
    }

    const tokens = [];
    cols.forEach((c) => {
      const cc = clean(c);
      if (!cc) return;
      tokens.push(cc);
      cc.split('|').map(clean).filter(Boolean).forEach((x) => tokens.push(x));
    });

    for (const token of tokens) {
      const up = token.toUpperCase();
      
      // Extract country from labeled format "Страна происхождения: XXX (code)"
      const countryLabel = token.match(/Страна происхождения\s*:\s*([^|]+)/i);
      if (!rec.country && countryLabel) {
        rec.country = shortCountry(countryLabel[1]);
      }
      
      // Extract country from format with code like "РОССИЯ (643)" or "АВСТРИЯ (040)"
      if (!rec.country && /\(\d{3}\)/.test(token)) {
        // Skip tokens that look like product names (start with digit + dot)
        if (!/^\d+\./.test(clean(token))) {
          rec.country = shortCountry(token);
        }
      }
      
      if (!rec.okpd2 && rx.okpd2.test(token)) {
        rec.okpd2 = clean((token.match(rx.okpd2) || [])[1] || '');
        rec.category_ls = clean(token.replace(rx.okpd2, '').replace(/[()]/g, ''));
      }
      if (up === 'ТОВАР') continue;
      if (up.includes('НДС')) {
        if (!rec.sum_rub) {
          const sumMatch = token.match(/(\d{1,3}(?:\s\d{3})+,\d{2}|\d+,\d{2})/);
          if (sumMatch) rec.sum_rub = clean(sumMatch[1]);
        }
        // Extract VAT rate
        const ndsMatch = token.match(/НДС[:\s]*(\d+\s*%?)/i);
        if (ndsMatch && !rec.nds) {
          let ndsVal = clean(ndsMatch[1]);
          if (!ndsVal.includes('%')) ndsVal = ndsVal + '%';
          rec.nds = ndsVal;
        }
        continue;
      }
      if (!rec.qty_consumption_unit && looksLikeQty(token)) {
        rec.qty_consumption_unit = extractQty(token) || clean(token);
      }
      if (!rec.price_per_unit && (looksLikePrice(token) || /\d{1,6}[.,]\d{2,7}/.test(token))) {
        rec.price_per_unit = extractPrice(token);
      }
      // Sum is exported only when VAT label is explicitly present.
    }

    if (!rec.qty_consumption_unit) rec.qty_consumption_unit = extractQty(rowText);
    if (!rec.price_per_unit) rec.price_per_unit = extractPrice(rowText);
    if (!rec.sum_rub && /НДС/i.test(rowText)) {
      const sumMatch = rowText.match(/(\d{1,3}(?:\s\d{3})+,\d{2}|\d+,\d{2})/);
      if (sumMatch) rec.sum_rub = clean(sumMatch[1]);
      // Extract VAT rate from row text
      const ndsMatch = rowText.match(/НДС[:\s]*(\d+\s*%?)/i);
      if (ndsMatch && !rec.nds) {
        let ndsVal = clean(ndsMatch[1]);
        if (!ndsVal.includes('%')) ndsVal = ndsVal + '%';
        rec.nds = ndsVal;
      }
    }

    if (!rec.category_ls) {
      const c2 = clean(cols[1] || '');
      rec.okpd2 = rec.okpd2 || clean((c2.match(rx.okpd2) || [])[1] || '');
      rec.category_ls = clean(c2.replace(rx.okpd2, '').replace(/[()]/g, ''));
    }
    return rec;
  };

  const parseNestedTable = (table) => {
    const rows = Array.from(table.querySelectorAll('tr'));
    const hMap = {};
    let headerFound = false;

    for (const tr of rows) {
      const cells = Array.from(tr.querySelectorAll('th,td'));
      if (!cells.length) continue;
      const headers = cells.map((x) => text(x).toUpperCase());
      const headerLine = headers.join(' | ');
      if (!/ТОРГОВОЕ НАИМЕНОВАНИЕ/.test(headerLine) || !/НОМЕР РУ/.test(headerLine)) continue;
      
      // Find column indices by matching header text explicitly
      // Note: first data row has chevron cell at index 0, so data columns are shifted by +1 relative to headers
      for (let i = 0; i < headers.length; i++) {
        const h = headers[i];
        // Match exact header names to avoid confusion with similar labels
        if (/ТОРГОВОЕ\s+НАИМЕНОВАНИЕ/.test(h)) hMap.trade = i;
        if (/НОМЕР\s+РУ/.test(h)) hMap.ru = i;
        if (/ЛЕКАРСТВЕННАЯ\s+ФОРМА/.test(h)) hMap.form = i;
        if (/ДОЗИРОВКА/.test(h)) hMap.dose = i;
        if (/КОЛИЧЕСТВО В ПОТРЕБ\.? ЕДИНИЦЕ/i.test(h)) hMap.qty = i;
      }
      headerFound = true;
      break;
    }

    if (!headerFound) return [];

    const details = [];
    for (let ri = 0; ri < rows.length; ri++) {
      const tr = rows[ri];
      const cols = Array.from(tr.querySelectorAll('td')).map((td) => text(td));
      if (!cols.length) continue;
      const rowText = clean(text(tr));
      const rowUp = rowText.toUpperCase();
      if (rowUp.includes('ТОРГОВОЕ НАИМЕНОВАНИЕ') || rowUp.includes('НОМЕР РУ')) continue;

      // Check if this is a detail row (contractSubjectDrugInfo) - skip for main data but extract meta
      const isDetailRow = tr.className && tr.className.includes('contractSubjectDrugInfo');
      
      if (isDetailRow) {
        // This is a detail row - extract holder, manufacturer, etc. from sections
        // Look for the last added detail record and update it
        if (details.length > 0) {
          const lastRec = details[details.length - 1];
          const sections = tr.querySelectorAll('.blockInfo__section');
          sections.forEach(sec => {
            const titleEl = sec.querySelector('.section__title');
            const infoEl = sec.querySelector('.section__info');
            if (!titleEl || !infoEl) return;
            const title = clean(text(titleEl));
            const info = clean(text(infoEl));
            if (!info) return;
            
            const titleUp = title.toUpperCase();
            if (/МЕЖДУНАРОДНОЕ.*НАИМЕНОВАНИЕ/.test(titleUp)) {
              if (!lastRec.mnn) lastRec.mnn = info;
            } else if (titleUp.includes('ДЕРЖАТЕЛЯ') || titleUp.includes('ВЛАДЕЛЕЦ')) {
              if (!lastRec.holder_name) lastRec.holder_name = info;
            } else if (titleUp.includes('ПРОИЗВОДИТЕЛЯ')) {
              if (!lastRec.manufacturer_name) lastRec.manufacturer_name = info;
            } else if (titleUp.includes('СТРАНА ПРОИЗВОДИТЕЛЯ')) {
              if (!lastRec.manufacturer_country) lastRec.manufacturer_country = info;
            } else if (titleUp.includes('ВИД ПЕРВИЧНОЙ УПАКОВКИ')) {
              if (!lastRec.primary_package_type) lastRec.primary_package_type = info;
            } else if (titleUp.includes('КОЛИЧЕСТВО ЛЕКАРСТВЕННЫХ ФОРМ')) {
              if (!lastRec.qty_forms_primary) lastRec.qty_forms_primary = info;
            } else if (titleUp.includes('КОЛИЧЕСТВО ПЕРВИЧНЫХ УПАКОВОК')) {
              if (!lastRec.qty_primary_packages) lastRec.qty_primary_packages = info;
            } else if (titleUp.includes('КОЛИЧЕСТВО ПОТРЕБИТЕЛЬСКИХ ЕДИНИЦ')) {
              if (!lastRec.qty_consumer_units) lastRec.qty_consumer_units = info;
            } else if (titleUp.includes('КОМПЛЕКТНОСТЬ')) {
              if (!lastRec.consumer_package_completeness) lastRec.consumer_package_completeness = info;
            }
          });
        }
        continue;
      }

      // Check for MNN in the expanded block sections (not just contractSubjectDrugInfo rows)
      // This code runs in the context of parseNestedTable, but we need to capture MNN/country from expanded sections
      const allSections = tr.querySelectorAll('.blockInfo__section');
      allSections.forEach(sec => {
        const titleEl = sec.querySelector('.section__title');
        const infoEl = sec.querySelector('.section__info');
        if (!titleEl || !infoEl) return;
        const title = clean(text(titleEl));
        const info = clean(text(infoEl));
        if (!info) return;

        const titleUp = title.toUpperCase();
        if (/МЕЖДУНАРОДНОЕ.*НАИМЕНОВАНИЕ/.test(titleUp)) {
          // Store in a temporary variable to be merged later
          window._tempMnn = window._tempMnn || info;
        } else if (titleUp.includes('СТРАНА ПРОИЗВОДИТЕЛЯ')) {
          window._tempManufacturerCountry = window._tempManufacturerCountry || info;
        }
      });

      // Skip rows that don't have the chevron cell (not data rows)
      const firstTd = tr.querySelector('td');
      const hasChevron = cols.length > 0 && (cols[0] === '' || (firstTd && firstTd.className && firstTd.className.includes('chevronRight')));
      if (!hasChevron) continue;

      const rec = blank();
      
      // Data columns start at index 1 (index 0 is chevron)
      // Map header indices to data indices: data_index = header_index + 1
      // But since we're reading from td elements only, and header had th elements,
      // the mapping is direct: cols[1] = trade, cols[2] = ru, cols[3] = form, cols[4] = dose, cols[5] = qty
      const tradeRaw = cols[1] || '';
      const ruRaw = cols[2] || '';
      const formRaw = cols[3] || '';
      const doseRaw = cols[4] || '';
      const qtyRaw = cols[5] || '';

      // Skip cells that contain meta blob (МНН и форма выпуска в соответствии с ГРЛС...)
      rec.trade_name = /МНН\s*:/i.test(tradeRaw) ? '' : clean(tradeRaw);
      
      // Fix RU extraction: take full text from TD, not just regex match
      rec.ru = clean(ruRaw);
      
      rec.release_form = clean(formRaw);
      
      // Dose may be split across spans - take full text
      rec.dose = clean(doseRaw);
      
      // Quantity: try column 5 first, fallback to regex search
      let qtyVal = clean(qtyRaw);
      if (!qtyVal || qtyVal.length < 1) {
        // Fallback: search for number in the cell
        const numMatch = qtyRaw.match(/\d[\d\s]*/);
        if (numMatch) qtyVal = clean(numMatch[0]);
      }
      if (qtyVal) rec.qty_consumption_unit = qtyVal;

      if (Object.values(rec).some(Boolean)) details.push(rec);
    }
    return details;
  };

  const nestedTables = Array.from(section.querySelectorAll('table')).filter((t) => {
    const tText = text(t).toUpperCase();
    return tText.includes('ТОРГОВОЕ НАИМЕНОВАНИЕ') && tText.includes('НОМЕР РУ');
  });

  const topRecords = topRows.map(parseTopRow);
  const detailsByIndex = nestedTables.map(parseNestedTable);
  const out = [];

  for (let i = 0; i < topRecords.length; i++) {
    const top = topRecords[i];
    const details = detailsByIndex[i] || [];
    if (!details.length) {
      // Will be handled in the second pass for expanded sections
      continue;
    }

    for (const d of details) {
      const rec = {
        ...top,
        mnn: clean(d.mnn || top.mnn),
        trade_name: clean(d.trade_name || top.trade_name),
        ru: clean(d.ru || top.ru),
        release_form: clean(d.release_form || top.release_form),
        dose: clean(d.dose || top.dose),
        qty_consumption_unit: clean(d.qty_consumption_unit || top.qty_consumption_unit),
        price_per_unit: clean(d.price_per_unit || top.price_per_unit),
        sum_rub: clean(d.sum_rub || top.sum_rub),
        nds: clean(d.nds || top.nds),
        holder_name: clean(d.holder_name || top.holder_name),
        manufacturer_name: clean(d.manufacturer_name || top.manufacturer_name),
        manufacturer_country: clean(d.manufacturer_country || top.manufacturer_country),
        primary_package_type: clean(d.primary_package_type || top.primary_package_type),
        qty_forms_primary: clean(d.qty_forms_primary || top.qty_forms_primary),
        qty_primary_packages: clean(d.qty_primary_packages || top.qty_primary_packages),
        qty_consumer_units: clean(d.qty_consumer_units || top.qty_consumer_units),
        consumer_package_completeness: clean(d.consumer_package_completeness || top.consumer_package_completeness),
      };
      out.push(rec);
    }
  }

  // Handle case where top row has no nested details but still has MNN/manufacturer_country in expanded sections
  for (let i = 0; i < topRecords.length; i++) {
    const top = topRecords[i];
    const details = detailsByIndex[i] || [];
    if (details.length > 0) continue; // Already handled above

    // Look for expanded sections directly after the top row
    const tr = topRows[i];
    if (!tr) continue;

    // Find next sibling row with ktru-tr-XXXXX id pattern that might contain expanded data
    let nextRow = tr.nextElementSibling;
    while (nextRow && nextRow.tagName === 'TR') {
      const rowId = nextRow.id || '';
      if (rowId.startsWith('ktru-tr-')) {
        // This is an expanded detail row - extract MNN and manufacturer country
        const sections = nextRow.querySelectorAll('.blockInfo__section');
        for (const sec of sections) {
          const titleEl = sec.querySelector('.section__title');
          const infoEl = sec.querySelector('.section__info');
          if (!titleEl || !infoEl) continue;
          const title = clean(text(titleEl));
          const info = clean(text(infoEl));
          if (!info) continue;

          const titleUp = title.toUpperCase();
          if (/МЕЖДУНАРОДНОЕ.*НАИМЕНОВАНИЕ/.test(titleUp)) {
            if (!top.mnn) top.mnn = info;
          } else if (titleUp.includes('СТРАНА ПРОИЗВОДИТЕЛЯ')) {
            if (!top.manufacturer_country) top.manufacturer_country = info;
          }
        }
        break;
      }
      nextRow = nextRow.nextElementSibling;
    }
    out.push(top);
  }

  // Each top row is a separate position - do not deduplicate by name alone.
  // Positions with same name but different qty/price are distinct entries.
  return out;
}
"""


if __name__ == "__main__":
    raise SystemExit(main())